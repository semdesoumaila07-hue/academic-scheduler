"""
Exporteur Excel pour les données du système.

Génère des fichiers Excel pour l'analyse et l'archivage.
"""
from typing import List, Dict
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from ..database.models import (
    ScheduleSlotModel, AcademicActivityModel, TeacherModel,
    CohortModel, LeaveRequestModel
)


class ExcelExporter:
    """
    Exporteur pour générer des fichiers Excel.
    
    Formats supportés :
    - Emploi du temps
    - Liste des activités
    - Liste des enseignants
    - Demandes de congés
    - Rapports de retards
    """
    
    def __init__(self, output_dir: Path = None):
        """
        Initialise l'exporteur Excel.
        
        Args:
            output_dir: Répertoire de sortie (outputs/exports par défaut)
        """
        if output_dir is None:
            output_dir = Path("outputs/exports")
        
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_schedule(self, cohort: CohortModel, slots: List[ScheduleSlotModel]) -> Path:
        """
        Exporte un emploi du temps en Excel.
        
        Args:
            cohort: Cohorte
            slots: Créneaux horaires
            
        Returns:
            Chemin du fichier Excel généré
        """
        filename = f"emploi_du_temps_{cohort.name.replace(' ', '_')}_{date.today()}.xlsx"
        filepath = self.output_dir / filename
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Emploi du Temps"
        
        # En-têtes
        headers = ['Date', 'Jour', 'Heure Début', 'Heure Fin', 'Activité', 
                  'Type', 'Enseignant', 'Salle', 'Retard (h)']
        ws.append(headers)
        
        # Style des en-têtes
        header_fill = PatternFill(start_color="34495e", end_color="34495e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Données
        for slot in sorted(slots, key=lambda s: (s.date, s.start_time)):
            row = [
                slot.date.strftime('%d/%m/%Y'),
                self._get_day_name(slot.date.weekday()),
                slot.start_time.strftime('%H:%M'),
                slot.end_time.strftime('%H:%M'),
                slot.activity.name if slot.activity else 'N/A',
                slot.activity.type.value if slot.activity else 'N/A',
                slot.teacher.full_name if slot.teacher else 'N/A',
                slot.room or 'N/A',
                f"{slot.delay_value:.2f}" if slot.delay_value else '0.00'
            ]
            ws.append(row)
        
        # Ajuster les largeurs de colonnes
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Bordures
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # Sauvegarder
        wb.save(filepath)
        
        return filepath
    
    def export_activities(self, activities: List[AcademicActivityModel]) -> Path:
        """
        Exporte la liste des activités en Excel.
        
        Args:
            activities: Liste des activités
            
        Returns:
            Chemin du fichier Excel généré
        """
        filename = f"activites_{date.today()}.xlsx"
        filepath = self.output_dir / filename
        
        # Préparer les données
        data = []
        for activity in activities:
            data.append({
                'Code': activity.code,
                'Nom': activity.name,
                'Type': activity.type.value,
                'Volume (h)': activity.volume_hours,
                'Réalisé (h)': activity.hours_done,
                'Restant (h)': activity.volume_hours - activity.hours_done,
                'Complétion (%)': round((activity.hours_done / activity.volume_hours * 100) if activity.volume_hours > 0 else 0, 1),
                'Facteur U': f"{activity.charge_factor:.3f}",
                'Priorité': activity.priority,
                'Statut': activity.status.value,
                'Date activation': activity.activation_date.strftime('%d/%m/%Y') if activity.activation_date else '',
                'Deadline': activity.deadline.strftime('%d/%m/%Y') if activity.deadline else ''
            })
        
        # Créer DataFrame
        df = pd.DataFrame(data)
        
        # Exporter
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Activités', index=False)
            
            # Formater
            workbook = writer.book
            worksheet = writer.sheets['Activités']
            
            # En-têtes
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        return filepath
    
    def export_teachers(self, teachers: List[TeacherModel]) -> Path:
        """
        Exporte la liste des enseignants en Excel.
        
        Args:
            teachers: Liste des enseignants
            
        Returns:
            Chemin du fichier Excel généré
        """
        filename = f"enseignants_{date.today()}.xlsx"
        filepath = self.output_dir / filename
        
        data = []
        for teacher in teachers:
            data.append({
                'Nom complet': teacher.full_name,
                'Email': teacher.email,
                'Téléphone': teacher.phone or '',
                'Spécialité': teacher.speciality,
                'Statut': teacher.status.value,
                'Max h/semaine': teacher.max_hours_per_week,
                'Max h/jour': teacher.max_hours_per_day
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Enseignants', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Enseignants']
            
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        return filepath
    
    def export_leaves(self, leaves: List[LeaveRequestModel]) -> Path:
        """
        Exporte les demandes de congés en Excel.
        
        Args:
            leaves: Liste des demandes de congés
            
        Returns:
            Chemin du fichier Excel généré
        """
        filename = f"conges_{date.today()}.xlsx"
        filepath = self.output_dir / filename
        
        data = []
        for leave in leaves:
            data.append({
                'Enseignant': leave.teacher.full_name if leave.teacher else 'N/A',
                'Type': leave.leave_type.value,
                'Début': leave.start_date.strftime('%d/%m/%Y'),
                'Fin': leave.end_date.strftime('%d/%m/%Y'),
                'Jours ouvrables': leave.working_days or 0,
                'Statut': leave.status.value,
                'Raison': leave.reason[:50] + '...' if len(leave.reason) > 50 else leave.reason,
                'Demande le': leave.created_at.strftime('%d/%m/%Y') if leave.created_at else '',
                'Approuvé par': leave.approver_email or '',
                'Approuvé le': leave.approved_at.strftime('%d/%m/%Y') if leave.approved_at else ''
            })
        
        df = pd.DataFrame(data)
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Congés', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Congés']
            
            for cell in worksheet[1]:
                cell.fill = PatternFill(start_color="e67e22", end_color="e67e22", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
        
        return filepath
    
    def export_delays_report(self, cohort: CohortModel, activities_delays: List[Dict]) -> Path:
        """
        Exporte un rapport de retards en Excel.
        
        Args:
            cohort: Cohorte
            activities_delays: Informations sur les retards
            
        Returns:
            Chemin du fichier Excel généré
        """
        filename = f"rapport_retards_{cohort.name.replace(' ', '_')}_{date.today()}.xlsx"
        filepath = self.output_dir / filename
        
        wb = Workbook()
        
        # Feuille 1 : Résumé
        ws_summary = wb.active
        ws_summary.title = "Résumé"
        
        ws_summary['A1'] = f"Rapport de Retards - {cohort.name}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Année académique:"
        ws_summary['B3'] = cohort.academic_year
        ws_summary['A4'] = "Semestre:"
        ws_summary['B4'] = cohort.semester
        ws_summary['A5'] = "Date du rapport:"
        ws_summary['B5'] = date.today().strftime('%d/%m/%Y')
        
        total_delay = sum(a.get('delay', 0) for a in activities_delays if a.get('delay', 0) > 0)
        ws_summary['A7'] = "Retard total:"
        ws_summary['B7'] = f"{total_delay:.1f} heures"
        ws_summary['B7'].font = Font(color="e74c3c", bold=True)
        
        # Feuille 2 : Détails
        ws_details = wb.create_sheet("Détails")
        
        headers = ['Activité', 'Code', 'Volume (h)', 'Réalisé (h)', 'Restant (h)', 
                  'Retard (h)', 'α', 'Urgence', 'Complétion (%)']
        ws_details.append(headers)
        
        # Style en-têtes
        for cell in ws_details[1]:
            cell.fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        for info in activities_delays:
            row = [
                info.get('activity_name', 'N/A'),
                info.get('activity_id', ''),
                info.get('volume_hours', 0),
                info.get('hours_done', 0),
                info.get('remaining_hours', 0),
                round(info.get('delay', 0), 2),
                round(info.get('alpha', 0), 3),
                info.get('urgency', 'Normal'),
                round(info.get('completion', 0), 1)
            ]
            ws_details.append(row)
        
        wb.save(filepath)
        
        return filepath
    
    @staticmethod
    def _get_day_name(weekday: int) -> str:
        """Retourne le nom du jour en français."""
        days = ['Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'Dimanche']
        return days[weekday]