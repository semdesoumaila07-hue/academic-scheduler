"""
<<<<<<< HEAD
Onglet de gestion des activités académiques - VERSION CORRIGÉE SQLite
Corrections :
  1. Import DataManager obsolète supprimé → db_manager + ActivityManager
  2. volume_spin.setValue() et realise_spin.setValue() → int() cast (fix TypeError float)
  3. edit_activity_by_index → mise à jour SQLite réelle (plus en mémoire)
  4. delete_activity_by_index → suppression SQLite réelle (plus en mémoire)
=======
Onglet de gestion des activités académiques - VERSION CORRIGÉE
Toutes les fonctionnalités sont maintenant connectées
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLabel, QLineEdit, QComboBox,
    QProgressBar, QDialog, QFormLayout, QSpinBox, QTextEdit,
<<<<<<< HEAD
    QMessageBox, QFileDialog, QDateEdit
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor
from datetime import datetime
import json

# ✅ Import corrigé — plus de DataManager
from src.database.db_manager import db_manager
from src.managers.activity_manager import ActivityManager
from src.database.repositories import CohortRepository, TeacherRepository


# ══════════════════════════════════════════════════════════
# CONSTANTES PRIORITÉS
# ══════════════════════════════════════════════════════════

PRIORITES = ["Urgente", "Possible", "Interdite"]

PRIORITE_STYLE = {
    "Urgente":  {"bg": "#FFCDD2", "fg": "#C62828", "icon": "🔴"},
    "Possible": {"bg": "#C8E6C9", "fg": "#2E7D32", "icon": "🟢"},
    "Interdite":{"bg": "#FFE0B2", "fg": "#E65100", "icon": "🟠"},
}


# ══════════════════════════════════════════════════════════
# DIALOGUE ACTIVITÉ
# ══════════════════════════════════════════════════════════

class ActivityDialog(QDialog):
    """Dialogue pour ajouter/modifier une activité académique."""

    def __init__(self, parent=None, activity_data=None,
                 cohortes=None, enseignants=None):
        super().__init__(parent)
        self.activity_data = activity_data or {}
        self.cohortes      = cohortes or []
        self.enseignants   = enseignants or []
        self.edit_mode     = activity_data is not None

        self.setWindowTitle(
            "Modifier Activité" if self.edit_mode else "Nouvelle Activité"
        )
        self.setModal(True)
        self.setMinimumWidth(620)
        self.setStyleSheet("background-color: white;")
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(5)

        title = QLabel(
            "✏️ Modifier Activité" if self.edit_mode
            else "➕ Nouvelle Activité Académique"
        )
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1a1a1a;")
        layout.addWidget(title)
        layout.addSpacing(15)

        form = QFormLayout()
        form.setSpacing(14)
        form.setLabelAlignment(Qt.AlignRight)

=======
    QMessageBox, QFileDialog, QCheckBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor
from datetime import datetime, timedelta
import json


class ActivityDialog(QDialog):
    """Dialogue pour ajouter/modifier une activité académique"""
    
    def __init__(self, parent=None, activity_data=None, cohortes=None, enseignants=None):
        super().__init__(parent)
        self.activity_data = activity_data or {}
        self.cohortes = cohortes or []
        self.enseignants = enseignants or []
        self.edit_mode = activity_data is not None
        
        self.setWindowTitle("Modifier Activité" if self.edit_mode else "Nouvelle Activité")
        self.setModal(True)
        self.setMinimumWidth(600)
        self.init_ui()
    
    def init_ui(self):
        """Initialiser l'interface du dialogue"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Titre
        title = QLabel("Modifier Activité" if self.edit_mode else "Nouvelle Activité Académique")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #1a1a1a;")
        layout.addWidget(title)
        layout.addSpacing(20)
        
        # Formulaire
        form = QFormLayout()
        form.setSpacing(15)
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        # Code
        self.code_edit = QLineEdit(self.activity_data.get('code', ''))
        self.code_edit.setPlaceholderText("Ex: ALGO-301")
        self.code_edit.setFixedHeight(40)
<<<<<<< HEAD
        self.code_edit.setStyleSheet(self._field_style())
        form.addRow("Code *:", self.code_edit)

=======
        form.addRow("Code *:", self.code_edit)
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        # Nom
        self.nom_edit = QLineEdit(self.activity_data.get('nom', ''))
        self.nom_edit.setPlaceholderText("Ex: Algorithmique avancée")
        self.nom_edit.setFixedHeight(40)
<<<<<<< HEAD
        self.nom_edit.setStyleSheet(self._field_style())
        form.addRow("Nom *:", self.nom_edit)

        # Type
        self.type_combo = QComboBox()
        self.type_combo.addItems(["CM", "TD", "TP", "Examen", "Soutenance"])
        self.type_combo.setFixedHeight(40)
        self.type_combo.setStyleSheet(self._field_style())
=======
        form.addRow("Nom *:", self.nom_edit)
        
        # Type
        self.type_combo = QComboBox()
        self.type_combo.addItems(["CM", "TD", "TP", "Examen", "Projet"])
        self.type_combo.setFixedHeight(40)
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        if 'type' in self.activity_data:
            idx = self.type_combo.findText(self.activity_data['type'])
            if idx >= 0:
                self.type_combo.setCurrentIndex(idx)
        form.addRow("Type *:", self.type_combo)
<<<<<<< HEAD

        # ✅ Volume horaire — int() pour éviter TypeError si float venant de SQLite
        self.volume_spin = QSpinBox()
        self.volume_spin.setRange(1, 200)
        self.volume_spin.setValue(int(self.activity_data.get('volume_heures', 20) or 20))
        self.volume_spin.setSuffix(" heures")
        self.volume_spin.setFixedHeight(40)
        self.volume_spin.setStyleSheet(self._field_style())
        form.addRow("Volume horaire *:", self.volume_spin)

        # ✅ Heures réalisées — int() pour éviter TypeError si float
        self.realise_spin = QSpinBox()
        self.realise_spin.setRange(0, 200)
        self.realise_spin.setValue(int(self.activity_data.get('heures_realisees', 0) or 0))
        self.realise_spin.setSuffix(" heures")
        self.realise_spin.setFixedHeight(40)
        self.realise_spin.setStyleSheet(self._field_style())
        form.addRow("Heures réalisées:", self.realise_spin)

        # Date d'activation
        self.activation_date_edit = QDateEdit()
        self.activation_date_edit.setCalendarPopup(True)
        self.activation_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.activation_date_edit.setFixedHeight(40)
        self.activation_date_edit.setStyleSheet(self._field_style())
        if 'activation_date' in self.activity_data:
            try:
                d = datetime.strptime(self.activity_data['activation_date'], "%d/%m/%Y")
                self.activation_date_edit.setDate(QDate(d.year, d.month, d.day))
            except Exception:
                self.activation_date_edit.setDate(QDate.currentDate())
        else:
            self.activation_date_edit.setDate(QDate.currentDate())
        form.addRow("Date d'activation:", self.activation_date_edit)

        # Échéance
        self.deadline_date_edit = QDateEdit()
        self.deadline_date_edit.setCalendarPopup(True)
        self.deadline_date_edit.setDisplayFormat("dd/MM/yyyy")
        self.deadline_date_edit.setFixedHeight(40)
        self.deadline_date_edit.setStyleSheet(self._field_style())
        if 'deadline' in self.activity_data:
            try:
                d = datetime.strptime(self.activity_data['deadline'], "%d/%m/%Y")
                self.deadline_date_edit.setDate(QDate(d.year, d.month, d.day))
            except Exception:
                self.deadline_date_edit.setDate(QDate.currentDate())
        else:
            self.deadline_date_edit.setDate(QDate.currentDate())
        form.addRow("Échéance:", self.deadline_date_edit)

        # Cohorte
        self.cohorte_combo = QComboBox()
        self.cohorte_combo.setFixedHeight(40)
        self.cohorte_combo.setStyleSheet(self._field_style())
        if self.cohortes:
            for c in self.cohortes:
                self.cohorte_combo.addItem(c.get('nom', 'N/A'), userData=c.get('id'))
        else:
            self.cohorte_combo.addItem("Aucune cohorte disponible", userData=None)
=======
        
        # Volume horaire
        self.volume_spin = QSpinBox()
        self.volume_spin.setRange(1, 200)
        self.volume_spin.setValue(self.activity_data.get('volume_heures', 20))
        self.volume_spin.setSuffix(" heures")
        self.volume_spin.setFixedHeight(40)
        form.addRow("Volume horaire *:", self.volume_spin)
        
        # Heures réalisées
        self.realise_spin = QSpinBox()
        self.realise_spin.setRange(0, 200)
        self.realise_spin.setValue(self.activity_data.get('heures_realisees', 0))
        self.realise_spin.setSuffix(" heures")
        self.realise_spin.setFixedHeight(40)
        form.addRow("Heures réalisées:", self.realise_spin)
        
        # Cohorte
        self.cohorte_combo = QComboBox()
        if self.cohortes:
            self.cohorte_combo.addItems([f"{c.get('nom', 'N/A')}" for c in self.cohortes])
        else:
            self.cohorte_combo.addItem("Aucune cohorte disponible")
        self.cohorte_combo.setFixedHeight(40)
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        if 'cohorte' in self.activity_data:
            idx = self.cohorte_combo.findText(self.activity_data['cohorte'])
            if idx >= 0:
                self.cohorte_combo.setCurrentIndex(idx)
        form.addRow("Cohorte *:", self.cohorte_combo)
<<<<<<< HEAD

        # Enseignant
        self.enseignant_combo = QComboBox()
        self.enseignant_combo.setFixedHeight(40)
        self.enseignant_combo.setStyleSheet(self._field_style())
        if self.enseignants:
            for e in self.enseignants:
                display_name = f"{e.get('nom', '')} {e.get('prenom', '')}".strip()
                if not display_name:
                    display_name = e.get('full_name', 'Enseignant')
                self.enseignant_combo.addItem(display_name, userData=e.get('id'))
        else:
            self.enseignant_combo.addItem("Aucun enseignant disponible", userData=None)
=======
        
        # Enseignant
        self.enseignant_combo = QComboBox()
        if self.enseignants:
            self.enseignant_combo.addItems([f"{e.get('nom', '')} {e.get('prenom', '')}" for e in self.enseignants])
        else:
            self.enseignant_combo.addItem("Aucun enseignant disponible")
        self.enseignant_combo.setFixedHeight(40)
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        if 'enseignant' in self.activity_data:
            idx = self.enseignant_combo.findText(self.activity_data['enseignant'])
            if idx >= 0:
                self.enseignant_combo.setCurrentIndex(idx)
        form.addRow("Enseignant *:", self.enseignant_combo)
<<<<<<< HEAD

        # Priorité
        self.priorite_combo = QComboBox()
        self.priorite_combo.setFixedHeight(40)
        self.priorite_combo.setStyleSheet(self._field_style())
        for p in PRIORITES:
            icon = PRIORITE_STYLE[p]["icon"]
            self.priorite_combo.addItem(f"{icon}  {p}", userData=p)
        if 'priorite' in self.activity_data:
            raw = self.activity_data['priorite']
            for i in range(self.priorite_combo.count()):
                if self.priorite_combo.itemData(i) == raw:
                    self.priorite_combo.setCurrentIndex(i)
                    break
        self.priorite_combo.currentIndexChanged.connect(self._update_priorite_color)
        self._update_priorite_color()
        form.addRow("Priorité *:", self.priorite_combo)

=======
        
        # Priorité
        self.priorite_combo = QComboBox()
        self.priorite_combo.addItems(["Normale", "Haute", "Urgente"])
        self.priorite_combo.setFixedHeight(40)
        if 'priorite' in self.activity_data:
            idx = self.priorite_combo.findText(self.activity_data['priorite'])
            if idx >= 0:
                self.priorite_combo.setCurrentIndex(idx)
        form.addRow("Priorité:", self.priorite_combo)
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        # Description
        self.description_edit = QTextEdit()
        self.description_edit.setPlaceholderText("Description de l'activité...")
        self.description_edit.setMaximumHeight(80)
<<<<<<< HEAD
        self.description_edit.setStyleSheet("""
            QTextEdit {
                border: 1px solid #D1D5DB; border-radius: 6px;
                padding: 8px; font-size: 13px;
            }
        """)
        if 'description' in self.activity_data:
            self.description_edit.setPlainText(self.activity_data['description'])
        form.addRow("Description:", self.description_edit)

        layout.addLayout(form)
        layout.addSpacing(15)

        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        btn_cancel = QPushButton("Annuler")
        btn_cancel.setFixedSize(140, 42)
        btn_cancel.setStyleSheet("""
            QPushButton { background: #e0e0e0; color: #333; border: none; border-radius: 6px; font-size: 14px; }
            QPushButton:hover { background: #d0d0d0; }
        """)
        btn_cancel.clicked.connect(self.reject)

        btn_save = QPushButton("💾 Enregistrer")
        btn_save.setFixedSize(160, 42)
        btn_save.setStyleSheet("""
            QPushButton { background: #4CAF50; color: white; border: none; border-radius: 6px; font-size: 14px; font-weight: bold; }
            QPushButton:hover { background: #45a049; }
        """)
        btn_save.clicked.connect(self.accept)

        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        layout.addLayout(btn_layout)

    def get_data(self):
        volume  = self.volume_spin.value()
        realise = self.realise_spin.value()
        progression = int((realise / volume) * 100) if volume > 0 else 0

=======
        if 'description' in self.activity_data:
            self.description_edit.setPlainText(self.activity_data['description'])
        form.addRow("Description:", self.description_edit)
        
        layout.addLayout(form)
        layout.addSpacing(10)
        
        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_cancel = QPushButton("❌ Annuler")
        btn_cancel.setFixedSize(140, 40)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                color: #333;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        btn_save = QPushButton("💾 Enregistrer")
        btn_save.setFixedSize(160, 40)
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_save.clicked.connect(self.accept)
        
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        
        layout.addLayout(btn_layout)
    
    def get_data(self):
        """Récupérer les données du formulaire"""
        volume = self.volume_spin.value()
        realise = self.realise_spin.value()
        progression = int((realise / volume) * 100) if volume > 0 else 0
        
        # Déterminer le statut
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        if progression == 0:
            statut = "En attente"
        elif progression == 100:
            statut = "Terminée"
<<<<<<< HEAD
        else:
            statut = "En cours"

        return {
            'code':             self.code_edit.text().strip(),
            'nom':              self.nom_edit.text().strip(),
            'type':             self.type_combo.currentText(),
            'volume_heures':    volume,
            'heures_realisees': realise,
            'progression':      progression,
            'cohorte_id':       self.cohorte_combo.currentData(),
            'cohorte':          self.cohorte_combo.currentText(),
            'teacher_id':       self.enseignant_combo.currentData(),
            'enseignant':       self.enseignant_combo.currentText(),
            'activation_date':  self.activation_date_edit.date().toString("dd/MM/yyyy"),
            'deadline':         self.deadline_date_edit.date().toString("dd/MM/yyyy"),
            'priorite':         self.priorite_combo.currentData(),
            'description':      self.description_edit.toPlainText().strip(),
            'statut':           statut,
        }

    def validate(self):
        data = self.get_data()
        if not data['code']:
            QMessageBox.warning(self, "Validation", "Le code est obligatoire !")
            return False
        if not data['nom']:
            QMessageBox.warning(self, "Validation", "Le nom est obligatoire !")
            return False
        if data['heures_realisees'] > data['volume_heures']:
            QMessageBox.warning(self, "Validation",
                "Les heures réalisées ne peuvent pas dépasser le volume horaire !")
            return False
        return True

    def accept(self):
        if self.validate():
            super().accept()

    def _field_style(self):
        return """
            QLineEdit, QComboBox, QSpinBox, QDateEdit {
                border: 1px solid #D1D5DB; border-radius: 6px;
                padding: 0 10px; font-size: 13px; background: white;
            }
            QLineEdit:focus, QComboBox:focus, QSpinBox:focus, QDateEdit:focus {
                border: 1px solid #3498db;
            }
        """

    def _update_priorite_color(self):
        raw = self.priorite_combo.currentData()
        style = PRIORITE_STYLE.get(raw, {})
        bg = style.get("bg", "white")
        fg = style.get("fg", "#333")
        self.priorite_combo.setStyleSheet(f"""
            QComboBox {{
                border: 1px solid #D1D5DB; border-radius: 6px;
                padding: 0 10px; font-size: 14px; font-weight: bold;
                background: {bg}; color: {fg};
            }}
            QComboBox::drop-down {{ border: none; }}
        """)


# ══════════════════════════════════════════════════════════
# ONGLET ACTIVITÉS
# ══════════════════════════════════════════════════════════

class ActivitiesTab(QWidget):
    """Onglet pour gérer les activités académiques — 100% SQLite."""

    def __init__(self, parent=None, current_user=None):
        super().__init__(parent)
        self.activities          = []
        self.filtered_activities = []
        self.current_user = current_user
        self.session = db_manager.get_session()
        self.activity_manager = ActivityManager(self.session)
        self.load_related_data()
        self.init_ui()
        self.refresh_activities_from_db()
        self.apply_filters()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)

        title = QLabel("Activités Académiques")
        title.setStyleSheet("font-size: 32px; font-weight: bold; color: #1a1a1a;")
        layout.addWidget(title)

        subtitle = QLabel("Gérez les cours, TD, TP et leur ordonnancement")
        subtitle.setStyleSheet("font-size: 16px; color: #666;")
        layout.addWidget(subtitle)

        legend_layout = QHBoxLayout()
        legend_layout.setSpacing(15)
        for p, s in PRIORITE_STYLE.items():
            lbl = QLabel(f"{s['icon']}  {p}")
            lbl.setStyleSheet(f"""
                QLabel {{
                    background: {s['bg']}; color: {s['fg']};
                    border-radius: 10px; padding: 4px 14px;
                    font-size: 13px; font-weight: bold;
                }}
            """)
            legend_layout.addWidget(lbl)
        legend_layout.addStretch()
        layout.addLayout(legend_layout)

        search_layout = QHBoxLayout()
=======
        elif progression > 0:
            statut = "En cours"
        else:
            statut = "Planifiée"
        
        return {
            'code': self.code_edit.text().strip(),
            'nom': self.nom_edit.text().strip(),
            'type': self.type_combo.currentText(),
            'volume_heures': volume,
            'heures_realisees': realise,
            'progression': progression,
            'cohorte': self.cohorte_combo.currentText(),
            'enseignant': self.enseignant_combo.currentText(),
            'priorite': self.priorite_combo.currentText(),
            'description': self.description_edit.toPlainText().strip(),
            'statut': statut
        }
    
    def validate(self):
        """Valider les données"""
        data = self.get_data()
        
        if not data['code']:
            QMessageBox.warning(self, "Validation", "Le code est obligatoire !")
            return False
        
        if not data['nom']:
            QMessageBox.warning(self, "Validation", "Le nom est obligatoire !")
            return False
        
        if data['heures_realisees'] > data['volume_heures']:
            QMessageBox.warning(self, "Validation", 
                              "Les heures réalisées ne peuvent pas dépasser le volume horaire !")
            return False
        
        return True
    
    def accept(self):
        """Accepter si validation OK"""
        if self.validate():
            super().accept()


class ActivitiesTab(QWidget):
    """Onglet pour gérer les activités académiques - VERSION CORRIGÉE"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.activities = []
        self.filtered_activities = []
        
        # Charger les données nécessaires
        self.load_related_data()
        
        self.init_ui()
        self.load_sample_data()
    
    def load_related_data(self):
        """Charger les cohortes et enseignants depuis les fichiers"""
        # Charger les cohortes
        try:
            import json
            from pathlib import Path
            
            # Cohortes
            structure_file = Path("data/structure.json")
            if structure_file.exists():
                with open(structure_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.cohortes = data.get('cohortes', [])
            else:
                self.cohortes = []
            
            # Enseignants (créer quelques exemples si pas de fichier)
            self.enseignants = [
                {'nom': 'KABORE', 'prenom': 'Marie'},
                {'nom': 'TRAORE', 'prenom': 'Moussa'},
                {'nom': 'SAWADOGO', 'prenom': 'Fatimata'},
                {'nom': 'OUATTARA', 'prenom': 'Ibrahim'},
                {'nom': 'ZONGO', 'prenom': 'Aminata'}
            ]
        except:
            self.cohortes = []
            self.enseignants = []
    
    def init_ui(self):
        """Initialise l'interface."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # En-tête
        header_layout = QVBoxLayout()
        title = QLabel("Activités Académiques")
        title.setStyleSheet("font-size: 32px; font-weight: bold; color: #1a1a1a;")
        header_layout.addWidget(title)
        
        subtitle = QLabel("Gérez les cours, TD, TP et leur ordonnancement")
        subtitle.setStyleSheet("font-size: 16px; color: #666;")
        header_layout.addWidget(subtitle)
        
        layout.addLayout(header_layout)
        
        # Barre de recherche et filtres
        search_layout = QHBoxLayout()
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍 Rechercher une activité...")
        self.search_box.setStyleSheet(self.get_input_style())
        self.search_box.setFixedHeight(45)
        self.search_box.textChanged.connect(self.apply_filters)
<<<<<<< HEAD

        self.filter_type = QComboBox()
        self.filter_type.addItems(["Tous types", "CM", "TD", "TP", "Examen", "Soutenance"])
        self.filter_type.setStyleSheet(self.get_input_style())
        self.filter_type.setFixedHeight(45)
        self.filter_type.currentTextChanged.connect(self.apply_filters)

        self.filter_status = QComboBox()
        self.filter_status.addItems(["Tous statuts", "En attente", "En cours", "Terminée"])
        self.filter_status.setStyleSheet(self.get_input_style())
        self.filter_status.setFixedHeight(45)
        self.filter_status.currentTextChanged.connect(self.apply_filters)

        self.filter_priorite = QComboBox()
        self.filter_priorite.addItems(["Toutes priorités"] + PRIORITES)
        self.filter_priorite.setStyleSheet(self.get_input_style())
        self.filter_priorite.setFixedHeight(45)
        self.filter_priorite.currentTextChanged.connect(self.apply_filters)

        search_layout.addWidget(self.search_box, 3)
        search_layout.addWidget(self.filter_type, 1)
        search_layout.addWidget(self.filter_status, 1)
        search_layout.addWidget(self.filter_priorite, 1)
        layout.addLayout(search_layout)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("➕ Nouvelle Activité")
        self.btn_add.setStyleSheet(self.get_button_style("#4CAF50"))
        self.btn_add.setFixedHeight(40)
        self.btn_add.clicked.connect(self.add_activity)

        self.btn_urgent = QPushButton("🔴 Activités Urgentes")
        self.btn_urgent.setStyleSheet(self.get_button_style("#F44336"))
        self.btn_urgent.setFixedHeight(40)
        self.btn_urgent.clicked.connect(self.show_urgent_activities)

        self.btn_delays = QPushButton("📊 Calculer Retards")
        self.btn_delays.setStyleSheet(self.get_button_style("#FF9800"))
        self.btn_delays.setFixedHeight(40)
        self.btn_delays.clicked.connect(self.calculate_delays)

        self.btn_export = QPushButton("📥 Exporter")
        self.btn_export.setStyleSheet(self.get_button_style("#2196F3"))
        self.btn_export.setFixedHeight(40)
        self.btn_export.clicked.connect(self.export_activities)

=======
        
        self.filter_type = QComboBox()
        self.filter_type.addItems(["Tous types", "CM", "TD", "TP", "Examen", "Projet"])
        self.filter_type.setStyleSheet(self.get_input_style())
        self.filter_type.setFixedHeight(45)
        self.filter_type.currentTextChanged.connect(self.apply_filters)
        
        self.filter_status = QComboBox()
        self.filter_status.addItems(["Tous statuts", "En attente", "Planifiée", "En cours", "Terminée"])
        self.filter_status.setStyleSheet(self.get_input_style())
        self.filter_status.setFixedHeight(45)
        self.filter_status.currentTextChanged.connect(self.apply_filters)
        
        search_layout.addWidget(self.search_box, 3)
        search_layout.addWidget(self.filter_type, 1)
        search_layout.addWidget(self.filter_status, 1)
        
        layout.addLayout(search_layout)
        
        # Boutons d'action
        btn_layout = QHBoxLayout()
        
        self.btn_add = QPushButton("➕ Nouvelle Activité")
        self.btn_add.setStyleSheet(self.get_button_style("#4CAF50"))
        self.btn_add.setFixedHeight(40)
        self.btn_add.clicked.connect(self.add_activity)  # ✅ CONNECTÉ
        
        self.btn_urgent = QPushButton("⚠️ Activités Urgentes")
        self.btn_urgent.setStyleSheet(self.get_button_style("#FF5722"))
        self.btn_urgent.setFixedHeight(40)
        self.btn_urgent.clicked.connect(self.show_urgent_activities)  # ✅ CONNECTÉ
        
        self.btn_delays = QPushButton("📊 Calculer Retards")
        self.btn_delays.setStyleSheet(self.get_button_style("#FF9800"))
        self.btn_delays.setFixedHeight(40)
        self.btn_delays.clicked.connect(self.calculate_delays)  # ✅ CONNECTÉ
        
        self.btn_export = QPushButton("📥 Exporter")
        self.btn_export.setStyleSheet(self.get_button_style("#2196F3"))
        self.btn_export.setFixedHeight(40)
        self.btn_export.clicked.connect(self.export_activities)  # ✅ CONNECTÉ
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_urgent)
        btn_layout.addWidget(self.btn_delays)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addStretch()
<<<<<<< HEAD
        layout.addLayout(btn_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(12)
        self.table.setHorizontalHeaderLabels([
            "Code", "Nom", "Type", "Volume (h)", "Réalisé (h)",
            "Progression", "Date d'activation", "Échéance",
            "Enseignant", "Cohorte", "Priorité", "Actions"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(11, QHeaderView.Fixed)
        self.table.setColumnWidth(11, 90)
=======
        
        layout.addLayout(btn_layout)
        
        # Table des activités
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "Code", "Nom", "Type", "Volume (h)", "Réalisé (h)", 
            "Progression", "Enseignant", "Cohorte", "Priorité", "Actions"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.table.setStyleSheet(self.get_table_style())
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
<<<<<<< HEAD
        layout.addWidget(self.table)

        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(15)
        self.stat_total    = self.create_stat_box("Total",        "0",  "#2196F3")
        self.stat_cm       = self.create_stat_box("CM",           "0",  "#3F51B5")
        self.stat_td       = self.create_stat_box("TD",           "0",  "#4CAF50")
        self.stat_tp       = self.create_stat_box("TP",           "0",  "#FF9800")
        self.stat_vol      = self.create_stat_box("Volume total", "0h", "#9C27B0")
        self.stat_realise  = self.create_stat_box("Réalisées",    "0h", "#009688")
        self.stat_urgentes = self.create_stat_box("Urgentes 🔴",  "0",  "#F44336")

        for s in [self.stat_total, self.stat_cm, self.stat_td, self.stat_tp,
                  self.stat_vol, self.stat_realise, self.stat_urgentes]:
            self.stats_layout.addWidget(s)
        layout.addLayout(self.stats_layout)

    # ── Chargement SQLite ─────────────────────────────────────────

    def load_related_data(self):
        """Charger les cohortes et enseignants depuis SQLite."""
        self.cohortes    = []
        self.enseignants = []
        try:
            cohort_rows = self.session.query(
                self.activity_manager.cohort_repo.model
            ).all()
            for c in cohort_rows:
                self.cohortes.append({
                    'id':  c.id,
                    'nom': f"{c.name} - {c.academic_year} S{c.semester}",
                    'academic_year': c.academic_year,
                    'semester': c.semester,
                })
            teacher_rows = self.session.query(
                self.activity_manager.teacher_repo.model
            ).all()
            for t in teacher_rows:
                self.enseignants.append({
                    'id':        t.id,
                    'nom':       '',
                    'prenom':    '',
                    'full_name': t.full_name,
                    'email':     t.email or '',
                })
        except Exception as e:
            print(f"❌ Erreur chargement données liées : {e}")

    def refresh_activities_from_db(self):
        """Recharger toutes les activités depuis SQLite."""
        self.activities = []
        try:
            db_activities = self.activity_manager.activity_repo.session.query(
                self.activity_manager.activity_repo.model
            ).all()
            for a in db_activities:
                self.activities.append({
                    'id':               a.id,
                    'code':             a.code,
                    'nom':              a.name,
                    'type':             a.type.value if hasattr(a.type, 'value') else str(a.type),
                    'volume_heures':    int(a.volume_hours or 0),
                    'heures_realisees': int(a.hours_done or 0),
                    'progression':      int((a.hours_done / a.volume_hours) * 100)
                                        if a.volume_hours and a.volume_hours > 0 else 0,
                    'cohorte':          self._get_cohort_name_by_id(a.cohort_id),
                    'cohorte_id':       a.cohort_id,
                    'enseignant':       self._get_teacher_name_by_id(a.teacher_id),
                    'teacher_id':       a.teacher_id,
                    'activation_date':  a.activation_date.strftime("%d/%m/%Y") if a.activation_date else "",
                    'deadline':         a.deadline.strftime("%d/%m/%Y") if a.deadline else "",
                    'priorite':         a.priority.value if hasattr(a.priority, 'value') else str(a.priority),
                    'description':      getattr(a, 'description', ""),
                    'statut':           a.status.value if hasattr(a.status, 'value') else str(a.status),
                })
        except Exception as e:
            print(f"❌ Erreur refresh activités : {e}")

    # ── CRUD ──────────────────────────────────────────────────────

    def add_activity(self):
        self.load_related_data()
        dialog = ActivityDialog(self, cohortes=self.cohortes, enseignants=self.enseignants)
        if dialog.exec_() == QDialog.Accepted:
            try:
                data = dialog.get_data()
                cohort_id  = data['cohorte_id']
                teacher_id = data.get('teacher_id')

                result = self.activity_manager.create_activity(
                    name=data['nom'],
                    code=data['code'],
                    activity_type=data['type'],
                    volume_hours=data['volume_heures'],
                    cohort_id=cohort_id,
                    teacher_id=teacher_id,
                    activation_date=self._parse_date(data['activation_date']),
                    deadline=self._parse_date(data['deadline']),
                    priority=data.get('priorite', None),
                    current_user=self.current_user
                )
                if result.get('success'):
                    self.refresh_activities_from_db()
                    self.apply_filters()
                    QMessageBox.information(self, "Succès", f"✅ Activité '{data['nom']}' ajoutée !")
                else:
                    QMessageBox.warning(self, "Erreur",
                        f"❌ Impossible d'ajouter :\n{result.get('error', 'Erreur inconnue')}")
            except Exception as e:
                import traceback; traceback.print_exc()
                QMessageBox.critical(self, "Erreur", f"❌ Erreur lors de l'ajout :\n{e}")

    def edit_activity_by_index(self, index):
        """✅ Modification SQLite réelle — plus en mémoire seulement."""
        if index >= len(self.filtered_activities):
            return
        activity = self.filtered_activities[index]
        self.load_related_data()
        dialog = ActivityDialog(self, activity, self.cohortes, self.enseignants)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            try:
                # Récupérer le modèle SQLAlchemy
                model = self.activity_manager.activity_repo.session.query(
                    self.activity_manager.activity_repo.model
                ).filter_by(id=activity['id']).first()

                if model:
                    model.name         = data['nom']
                    model.code         = data['code']
                    model.volume_hours = data['volume_heures']
                    model.hours_done   = data['heures_realisees']
                    model.cohort_id    = data['cohorte_id']
                    model.teacher_id   = data['teacher_id']
                    model.updated_at   = datetime.now()

                    # Activation date et deadline
                    if data.get('activation_date'):
                        model.activation_date = self._parse_date(data['activation_date'])
                    if data.get('deadline'):
                        model.deadline = self._parse_date(data['deadline'])

                    # Type → enum
                    try:
                        from src.database.models import ActivityTypeEnum
                        type_map = {e.value: e for e in ActivityTypeEnum}
                        if data['type'] in type_map:
                            model.type = type_map[data['type']]
                    except Exception:
                        pass

                    # Priorité → enum
                    try:
                        from src.database.models import ActivityPriorityEnum
                        prio_map = {e.value: e for e in ActivityPriorityEnum}
                        if data.get('priorite') in prio_map:
                            model.priority = prio_map[data['priorite']]
                    except Exception:
                        pass

                    self.session.commit()
                    self.refresh_activities_from_db()
                    self.apply_filters()
                    QMessageBox.information(self, "Succès",
                        f"✅ Activité '{data['nom']}' modifiée !")
                else:
                    QMessageBox.warning(self, "Erreur",
                        "Activité introuvable dans la base de données.")
            except Exception as e:
                self.session.rollback()
                import traceback; traceback.print_exc()
                QMessageBox.critical(self, "Erreur", f"❌ Erreur modification :\n{e}")

    def delete_activity_by_index(self, index):
        """✅ Suppression SQLite réelle — plus en mémoire seulement."""
        if index >= len(self.filtered_activities):
            return
        activity = self.filtered_activities[index]
        reply = QMessageBox.question(
            self, "Confirmation",
            f"Supprimer l'activité '{activity['nom']}' ?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                self.activity_manager.activity_repo.delete(activity['id'])
                self.session.commit()
                self.refresh_activities_from_db()
                self.apply_filters()
                QMessageBox.information(self, "Succès", "✅ Activité supprimée !")
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "Erreur", f"❌ Erreur suppression :\n{e}")

    # ── Filtres & affichage ───────────────────────────────────────

    def apply_filters(self):
        search_text    = self.search_box.text().lower()
        filter_type    = self.filter_type.currentText()
        filter_status  = self.filter_status.currentText()
        filter_priorite = self.filter_priorite.currentText()

        self.filtered_activities = []
        for activity in self.activities:
=======
        
        layout.addWidget(self.table)
        
        # Statistiques en bas
        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(15)
        
        self.stat1 = self.create_stat_box("Total activités", "0", "#2196F3")
        self.stat2 = self.create_stat_box("CM", "0", "#3F51B5")
        self.stat3 = self.create_stat_box("TD", "0", "#4CAF50")
        self.stat4 = self.create_stat_box("TP", "0", "#FF9800")
        self.stat5 = self.create_stat_box("Volume total", "0h", "#9C27B0")
        self.stat6 = self.create_stat_box("Réalisées", "0h", "#4CAF50")
        
        self.stats_layout.addWidget(self.stat1)
        self.stats_layout.addWidget(self.stat2)
        self.stats_layout.addWidget(self.stat3)
        self.stats_layout.addWidget(self.stat4)
        self.stats_layout.addWidget(self.stat5)
        self.stats_layout.addWidget(self.stat6)
        
        layout.addLayout(self.stats_layout)
    
    def load_sample_data(self):
        """Charger des données exemple"""
        self.activities = [
            {
                'id': 1,
                'code': 'ALGO-301',
                'nom': 'Algorithmique avancée',
                'type': 'CM',
                'volume_heures': 30,
                'heures_realisees': 30,
                'progression': 100,
                'enseignant': 'KABORE Marie',
                'cohorte': 'L3 Info',
                'priorite': 'Normale',
                'statut': 'Terminée',
                'description': 'Cours d\'algorithmique niveau avancé'
            },
            {
                'id': 2,
                'code': 'ALGO-TD-301',
                'nom': 'TD Algorithmique',
                'type': 'TD',
                'volume_heures': 20,
                'heures_realisees': 20,
                'progression': 100,
                'enseignant': 'KABORE Marie',
                'cohorte': 'L3 Info',
                'priorite': 'Normale',
                'statut': 'Terminée',
                'description': 'Travaux dirigés d\'algorithmique'
            },
            {
                'id': 3,
                'code': 'BD-301',
                'nom': 'Bases de données',
                'type': 'CM',
                'volume_heures': 25,
                'heures_realisees': 15,
                'progression': 60,
                'enseignant': 'TRAORE Moussa',
                'cohorte': 'L3 Info',
                'priorite': 'Haute',
                'statut': 'En cours',
                'description': 'Introduction aux bases de données relationnelles'
            },
            {
                'id': 4,
                'code': 'BD-TP-301',
                'nom': 'TP Bases de données',
                'type': 'TP',
                'volume_heures': 20,
                'heures_realisees': 10,
                'progression': 50,
                'enseignant': 'TRAORE Moussa',
                'cohorte': 'L3 Info',
                'priorite': 'Urgente',
                'statut': 'En cours',
                'description': 'Travaux pratiques sur PostgreSQL'
            },
            {
                'id': 5,
                'code': 'WEB-301',
                'nom': 'Développement Web',
                'type': 'CM',
                'volume_heures': 25,
                'heures_realisees': 0,
                'progression': 0,
                'enseignant': 'OUATTARA Ibrahim',
                'cohorte': 'L3 Info',
                'priorite': 'Urgente',
                'statut': 'En attente',
                'description': 'HTML, CSS, JavaScript'
            },
        ]
        self.apply_filters()
    
    def apply_filters(self):
        """Appliquer les filtres de recherche"""
        search_text = self.search_box.text().lower()
        filter_type = self.filter_type.currentText()
        filter_status = self.filter_status.currentText()
        
        self.filtered_activities = []
        
        for activity in self.activities:
            # Filtre de recherche
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            if search_text:
                searchable = f"{activity['code']} {activity['nom']} {activity['enseignant']} {activity['cohorte']}".lower()
                if search_text not in searchable:
                    continue
<<<<<<< HEAD
            if filter_type != "Tous types" and activity['type'] != filter_type:
                continue
            if filter_status != "Tous statuts" and activity['statut'] != filter_status:
                continue
            if filter_priorite != "Toutes priorités" and activity['priorite'] != filter_priorite:
                continue
            self.filtered_activities.append(activity)

        self.refresh_table()
        self.update_statistics()

    def refresh_table(self):
        self.table.setRowCount(len(self.filtered_activities))
        for i, activity in enumerate(self.filtered_activities):
            self.table.setRowHeight(i, 48)
            self.table.setItem(i, 0, QTableWidgetItem(activity['code']))
            self.table.setItem(i, 1, QTableWidgetItem(activity['nom']))

            type_item = QTableWidgetItem(activity['type'])
            type_colors = {
                "CM": "#E3F2FD", "TD": "#E8F5E9", "TP": "#FFF3E0",
                "Examen": "#F3E5F5", "Soutenance": "#FCE4EC"
            }
            type_item.setBackground(QColor(type_colors.get(activity['type'], "#FAFAFA")))
            self.table.setItem(i, 2, type_item)

            self.table.setItem(i, 3, QTableWidgetItem(f"{activity['volume_heures']}h"))
            self.table.setItem(i, 4, QTableWidgetItem(f"{activity['heures_realisees']}h"))

=======
            
            # Filtre de type
            if filter_type != "Tous types" and activity['type'] != filter_type:
                continue
            
            # Filtre de statut
            if filter_status != "Tous statuts" and activity['statut'] != filter_status:
                continue
            
            self.filtered_activities.append(activity)
        
        self.refresh_table()
        self.update_statistics()
    
    def refresh_table(self):
        """Rafraîchir l'affichage de la table"""
        self.table.setRowCount(len(self.filtered_activities))
        
        for i, activity in enumerate(self.filtered_activities):
            # Code
            self.table.setItem(i, 0, QTableWidgetItem(activity['code']))
            
            # Nom
            self.table.setItem(i, 1, QTableWidgetItem(activity['nom']))
            
            # Type
            type_item = QTableWidgetItem(activity['type'])
            if activity['type'] == "CM":
                type_item.setBackground(QColor("#E3F2FD"))
            elif activity['type'] == "TD":
                type_item.setBackground(QColor("#E8F5E9"))
            elif activity['type'] == "TP":
                type_item.setBackground(QColor("#FFF3E0"))
            self.table.setItem(i, 2, type_item)
            
            # Volume
            self.table.setItem(i, 3, QTableWidgetItem(f"{activity['volume_heures']}h"))
            
            # Réalisé
            self.table.setItem(i, 4, QTableWidgetItem(f"{activity['heures_realisees']}h"))
            
            # Progression (barre)
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            progress = QProgressBar()
            progress.setValue(activity['progression'])
            progress.setTextVisible(True)
            progress.setFormat(f"{activity['progression']}%")
<<<<<<< HEAD
            chunk_color = ("#4CAF50" if activity['progression'] == 100
                           else "#FF9800" if activity['progression'] > 50
                           else "#F44336")
            progress.setStyleSheet(f"""
                QProgressBar {{ border: 1px solid #E0E0E0; border-radius: 4px;
                    background: #F5F5F5; text-align: center; font-size: 12px; }}
                QProgressBar::chunk {{ background-color: {chunk_color}; border-radius: 4px; }}
            """)
            self.table.setCellWidget(i, 5, progress)

            self.table.setItem(i, 6, QTableWidgetItem(activity.get('activation_date', '')))
            self.table.setItem(i, 7, QTableWidgetItem(activity.get('deadline', '')))
            self.table.setItem(i, 8, QTableWidgetItem(activity['enseignant']))
            self.table.setItem(i, 9, QTableWidgetItem(activity['cohorte']))

            raw_p = activity.get('priorite', 'Possible')
            if raw_p in ("Normale", "Haute", "Basse"):
                raw_p = "Possible"
            elif raw_p in ("Innterdit", "Interdit"):
                raw_p = "Interdite"
            style_p = PRIORITE_STYLE.get(raw_p, PRIORITE_STYLE["Possible"])
            priorite_item = QTableWidgetItem(f"{style_p['icon']}  {raw_p}")
            priorite_item.setBackground(QColor(style_p["bg"]))
            priorite_item.setForeground(QColor(style_p["fg"]))
            self.table.setItem(i, 10, priorite_item)

            self.table.setCellWidget(i, 11, self._create_action_buttons(i))

    def _create_action_buttons(self, index):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(4, 0, 4, 0)
        layout.setSpacing(4)

        btn_edit = QPushButton("✏️")
        btn_edit.setFixedSize(32, 32)
        btn_edit.setToolTip("Modifier")
        btn_edit.setStyleSheet("""
            QPushButton { background: transparent; border: none; font-size: 15px; }
            QPushButton:hover { background: #E3F2FD; border-radius: 4px; }
        """)
        btn_edit.clicked.connect(lambda checked, idx=index: self.edit_activity_by_index(idx))

        btn_delete = QPushButton("🗑️")
        btn_delete.setFixedSize(32, 32)
        btn_delete.setToolTip("Supprimer")
        btn_delete.setStyleSheet("""
            QPushButton { background: transparent; border: none; font-size: 15px; }
            QPushButton:hover { background: #FFEBEE; border-radius: 4px; }
        """)
        btn_delete.clicked.connect(lambda checked, idx=index: self.delete_activity_by_index(idx))

        layout.addWidget(btn_edit)
        layout.addWidget(btn_delete)
        return widget

    # ── Statistiques ──────────────────────────────────────────────

    def update_statistics(self):
        acts = self.filtered_activities
        total    = len(acts)
        cm       = sum(1 for a in acts if a['type'] in ('CM', 'Cours Magistral', 'COURS_MAGISTRAL'))
        td       = sum(1 for a in acts if a['type'] in ('TD', 'Travaux Dirigés', 'Travaux Diriges', 'TD'))
        tp       = sum(1 for a in acts if a['type'] in ('TP', 'Travaux Pratiques', 'TP'))
        vol      = sum(a['volume_heures']    for a in acts)
        realise  = sum(a['heures_realisees'] for a in acts)
        urgentes = sum(1 for a in acts if a.get('priorite') == 'Urgente')

        self.update_stat_box(self.stat_total,    str(total))
        self.update_stat_box(self.stat_cm,       str(cm))
        self.update_stat_box(self.stat_td,       str(td))
        self.update_stat_box(self.stat_tp,       str(tp))
        self.update_stat_box(self.stat_vol,      f"{vol}h")
        self.update_stat_box(self.stat_realise,  f"{realise}h")
        self.update_stat_box(self.stat_urgentes, str(urgentes))

    def update_stat_box(self, box, value):
        labels = box.findChildren(QLabel)
        if labels:
            labels[0].setText(value)

    # ── Actions ───────────────────────────────────────────────────

    def show_urgent_activities(self):
        urgent = [a for a in self.activities if a.get('priorite') == 'Urgente']
        if not urgent:
            QMessageBox.information(self, "Activités Urgentes", "✅ Aucune activité urgente !")
            return
        msg = f"🔴 {len(urgent)} activité(s) urgente(s) :\n\n"
        for a in urgent:
            msg += f"• {a['code']} - {a['nom']}\n  Progression : {a['progression']}%\n\n"
        QMessageBox.warning(self, "Activités Urgentes", msg)
        self.filter_priorite.setCurrentText("Urgente")

    def calculate_delays(self):
        total_vol    = sum(a['volume_heures']    for a in self.activities)
        total_realise = sum(a['heures_realisees'] for a in self.activities)
        retard = total_vol - total_realise
        pct    = int(total_realise / total_vol * 100) if total_vol else 0
        en_retard = [a for a in self.activities
                     if a['progression'] < 100 and a['statut'] != 'En attente']
        msg = (f"📊 RAPPORT DE RETARDS\n\nVolume total : {total_vol}h\n"
               f"Réalisé : {total_realise}h\nRetard : {retard}h\nProgression : {pct}%\n\n")
        if en_retard:
            msg += f"⚠️ {len(en_retard)} activité(s) en retard :\n\n"
            for a in sorted(en_retard,
                            key=lambda x: x['volume_heures'] - x['heures_realisees'],
                            reverse=True)[:5]:
                r = a['volume_heures'] - a['heures_realisees']
                msg += f"• {a['nom']} — {r}h ({a['progression']}%)\n"
        else:
            msg += "✅ Aucune activité en retard !"
        QMessageBox.information(self, "Calcul des Retards", msg)

    def export_activities(self):
        reply = QMessageBox.question(self, "Format d'export",
            "Oui = Excel (.xlsx)\nNon = JSON (.json)",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
        if reply == QMessageBox.Yes:
            self._export_excel()
        else:
            self._export_json()

    def _export_excel(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, "Exporter en Excel",
            f"activites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)")
        if not filename:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = Workbook(); ws = wb.active; ws.title = "Activités"
            headers = ["Code", "Nom", "Type", "Volume (h)", "Réalisé (h)",
                       "Progression (%)", "Enseignant", "Cohorte", "Priorité", "Statut"]
            ws.append(headers)
=======
            
            if activity['progression'] == 100:
                progress.setStyleSheet("QProgressBar::chunk { background-color: #4CAF50; }")
            elif activity['progression'] > 50:
                progress.setStyleSheet("QProgressBar::chunk { background-color: #FF9800; }")
            else:
                progress.setStyleSheet("QProgressBar::chunk { background-color: #F44336; }")
            
            self.table.setCellWidget(i, 5, progress)
            
            # Enseignant
            self.table.setItem(i, 6, QTableWidgetItem(activity['enseignant']))
            
            # Cohorte
            self.table.setItem(i, 7, QTableWidgetItem(activity['cohorte']))
            
            # Priorité
            priorite_item = QTableWidgetItem(activity['priorite'])
            if activity['priorite'] == "Urgente":
                priorite_item.setBackground(QColor("#FFCDD2"))
            elif activity['priorite'] == "Haute":
                priorite_item.setBackground(QColor("#FFECB3"))
            self.table.setItem(i, 8, priorite_item)
            
            # Actions
            actions_widget = self.create_action_buttons(activity)
            self.table.setCellWidget(i, 9, actions_widget)
    
    def create_action_buttons(self, activity):
        """Créer les boutons d'action pour une ligne"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(5, 0, 5, 0)
        layout.setSpacing(5)
        
        btn_edit = QPushButton("✏️")
        btn_edit.setFixedSize(30, 30)
        btn_edit.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #E3F2FD;
                border-radius: 4px;
            }
        """)
        btn_edit.clicked.connect(lambda: self.edit_activity(activity))
        
        btn_delete = QPushButton("🗑️")
        btn_delete.setFixedSize(30, 30)
        btn_delete.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #FFEBEE;
                border-radius: 4px;
            }
        """)
        btn_delete.clicked.connect(lambda: self.delete_activity(activity))
        
        layout.addWidget(btn_edit)
        layout.addWidget(btn_delete)
        layout.addStretch()
        
        return widget
    
    def update_statistics(self):
        """Mettre à jour les statistiques"""
        total = len(self.filtered_activities)
        cm = sum(1 for a in self.filtered_activities if a['type'] == 'CM')
        td = sum(1 for a in self.filtered_activities if a['type'] == 'TD')
        tp = sum(1 for a in self.filtered_activities if a['type'] == 'TP')
        volume_total = sum(a['volume_heures'] for a in self.filtered_activities)
        realisees = sum(a['heures_realisees'] for a in self.filtered_activities)
        
        self.update_stat_box(self.stat1, str(total))
        self.update_stat_box(self.stat2, str(cm))
        self.update_stat_box(self.stat3, str(td))
        self.update_stat_box(self.stat4, str(tp))
        self.update_stat_box(self.stat5, f"{volume_total}h")
        self.update_stat_box(self.stat6, f"{realisees}h")
    
    def update_stat_box(self, box, value):
        """Mettre à jour une boîte de statistique"""
        labels = box.findChildren(QLabel)
        if labels:
            labels[0].setText(value)
    
    # ==========================================
    # MÉTHODES D'ACTION
    # ==========================================
    
    def add_activity(self):
        """Ajouter une nouvelle activité"""
        dialog = ActivityDialog(self, cohortes=self.cohortes, enseignants=self.enseignants)
        
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            
            # Générer un nouvel ID
            new_id = max([a['id'] for a in self.activities], default=0) + 1
            data['id'] = new_id
            
            # Ajouter à la liste
            self.activities.append(data)
            
            # Rafraîchir
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ L'activité '{data['nom']}' a été ajoutée avec succès !"
            )
    
    def edit_activity(self, activity):
        """Modifier une activité"""
        dialog = ActivityDialog(self, activity, self.cohortes, self.enseignants)
        
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            activity.update(data)
            
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ L'activité '{activity['nom']}' a été modifiée avec succès !"
            )
    
    def delete_activity(self, activity):
        """Supprimer une activité"""
        reply = QMessageBox.question(
            self,
            "Confirmation",
            f"Êtes-vous sûr de vouloir supprimer l'activité '{activity['nom']}' ?\n\n"
            "Cette action est irréversible.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.activities.remove(activity)
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ L'activité '{activity['nom']}' a été supprimée avec succès !"
            )
    
    def show_urgent_activities(self):
        """Afficher uniquement les activités urgentes"""
        urgent = [a for a in self.activities if a['priorite'] == 'Urgente']
        
        if not urgent:
            QMessageBox.information(
                self,
                "Activités Urgentes",
                "✅ Aucune activité urgente pour le moment !"
            )
            return
        
        # Créer un message détaillé
        message = f"⚠️ {len(urgent)} activité(s) urgente(s) détectée(s) :\n\n"
        
        for a in urgent:
            status = "🔴" if a['progression'] < 50 else "🟡"
            message += f"{status} {a['code']} - {a['nom']}\n"
            message += f"   Progression: {a['progression']}% | Enseignant: {a['enseignant']}\n\n"
        
        QMessageBox.warning(
            self,
            "Activités Urgentes",
            message
        )
        
        # Filtrer pour afficher seulement les urgentes
        self.filter_status.setCurrentText("Tous statuts")
        self.search_box.clear()
        self.filtered_activities = urgent
        self.refresh_table()
    
    def calculate_delays(self):
        """Calculer les retards des activités"""
        total_volume = sum(a['volume_heures'] for a in self.activities)
        total_realise = sum(a['heures_realisees'] for a in self.activities)
        retard_heures = total_volume - total_realise
        
        pourcentage_global = int((total_realise / total_volume) * 100) if total_volume > 0 else 0
        
        # Calculer les activités en retard
        en_retard = []
        for a in self.activities:
            if a['progression'] < 100 and a['statut'] != 'En attente':
                retard_act = a['volume_heures'] - a['heures_realisees']
                en_retard.append({
                    'nom': a['nom'],
                    'retard': retard_act,
                    'progression': a['progression']
                })
        
        # Créer le rapport
        message = "📊 RAPPORT DE RETARDS\n\n"
        message += f"Volume horaire total: {total_volume}h\n"
        message += f"Heures réalisées: {total_realise}h\n"
        message += f"Retard global: {retard_heures}h\n"
        message += f"Progression globale: {pourcentage_global}%\n\n"
        
        if en_retard:
            message += f"⚠️ {len(en_retard)} activité(s) en retard:\n\n"
            for item in sorted(en_retard, key=lambda x: x['retard'], reverse=True)[:5]:
                message += f"• {item['nom']}\n"
                message += f"  Retard: {item['retard']}h ({item['progression']}% complété)\n\n"
        else:
            message += "✅ Aucune activité en retard !"
        
        QMessageBox.information(
            self,
            "Calcul des Retards",
            message
        )
    
    def export_activities(self):
        """Exporter les activités"""
        reply = QMessageBox.question(
            self,
            "Format d'export",
            "Choisissez le format d'export :\n\n"
            "Oui = Excel (.xlsx)\n"
            "Non = JSON (.json)",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            self.export_to_excel()
        else:
            self.export_to_json()
    
    def export_to_excel(self):
        """Exporter en Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter en Excel",
            f"activites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Activités"
            
            # En-têtes
            headers = ["Code", "Nom", "Type", "Volume (h)", "Réalisé (h)", 
                      "Progression (%)", "Enseignant", "Cohorte", "Priorité", "Statut"]
            ws.append(headers)
            
            # Style des en-têtes
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
<<<<<<< HEAD
            for a in self.filtered_activities:
                ws.append([a['code'], a['nom'], a['type'], a['volume_heures'],
                           a['heures_realisees'], a['progression'],
                           a['enseignant'], a['cohorte'], a['priorite'], a['statut']])
            wb.save(filename)
            QMessageBox.information(self, "Succès", f"✅ Export Excel !\n{filename}")
        except ImportError:
            QMessageBox.warning(self, "Module manquant", "pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))

    def _export_json(self):
        filename, _ = QFileDialog.getSaveFileName(
            self, "Exporter en JSON",
            f"activites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON Files (*.json)")
        if not filename:
            return
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.filtered_activities, f, indent=2, ensure_ascii=False)
            QMessageBox.information(self, "Succès", f"✅ Export JSON !\n{filename}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))

    # ── Utilitaires ───────────────────────────────────────────────

    def _get_cohort_name_by_id(self, id):
        for c in self.cohortes:
            if c.get('id') == id:
                return c.get('nom', '')
        return ""

    def _get_teacher_name_by_id(self, id):
        for t in self.enseignants:
            if t.get('id') == id:
                return (f"{t.get('nom', '')} {t.get('prenom', '')}".strip()
                        or t.get('full_name', 'Enseignant'))
        return ""

    def _parse_date(self, date_str):
        try:
            return datetime.strptime(date_str, "%d/%m/%Y").date()
        except Exception:
            return None

    def create_stat_box(self, label, value, color):
        box = QWidget()
        box.setStyleSheet(f"QWidget {{ background-color: {color}; border-radius: 8px; padding: 12px; }}")
        layout = QVBoxLayout(box)
        v_lbl = QLabel(value)
        v_lbl.setStyleSheet("font-size: 20px; font-weight: bold; color: white;")
        v_lbl.setAlignment(Qt.AlignCenter)
        l_lbl = QLabel(label)
        l_lbl.setStyleSheet("font-size: 11px; color: white;")
        l_lbl.setAlignment(Qt.AlignCenter)
        layout.addWidget(v_lbl)
        layout.addWidget(l_lbl)
        return box

    def get_input_style(self):
        return """
            QLineEdit, QComboBox {
                padding: 10px; border: 1px solid #E0E0E0;
                border-radius: 8px; font-size: 14px; background: white;
            }
        """

    def get_button_style(self, color):
        return f"""
            QPushButton {{
                background-color: {color}; color: white; border: none;
                border-radius: 8px; padding: 10px 20px;
                font-size: 14px; font-weight: bold;
            }}
            QPushButton:hover {{ opacity: 0.9; }}
        """

    def get_table_style(self):
        return """
            QTableWidget { border: 1px solid #E0E0E0; border-radius: 8px;
                background-color: white; gridline-color: #F0F0F0; }
            QTableWidget::item { padding: 10px; }
            QTableWidget::item:selected { background-color: #E3F2FD; color: #1976D2; }
            QHeaderView::section { background-color: #F5F5F5; padding: 12px;
                border: none; font-weight: bold; color: #333; }
            QTableWidget::item:alternate { background-color: #FAFAFA; }
        """

    def refresh_data(self):
        self.load_related_data()
        self.refresh_activities_from_db()
        self.apply_filters()

    def showEvent(self, event):
        super().showEvent(event)
        self.refresh_data()
=======
            
            # Données
            for activity in self.filtered_activities:
                ws.append([
                    activity['code'],
                    activity['nom'],
                    activity['type'],
                    activity['volume_heures'],
                    activity['heures_realisees'],
                    activity['progression'],
                    activity['enseignant'],
                    activity['cohorte'],
                    activity['priorite'],
                    activity['statut']
                ])
            
            # Ajuster les colonnes
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ Export Excel réussi !\n\nFichier : {filename}\nActivités exportées : {len(self.filtered_activities)}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self,
                "Module manquant",
                "Le module 'openpyxl' n'est pas installé.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export Excel :\n\n{str(e)}"
            )
    
    def export_to_json(self):
        """Exporter en JSON"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter en JSON",
            f"activites_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            "JSON Files (*.json)"
        )
        
        if not filename:
            return
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.filtered_activities, f, indent=2, ensure_ascii=False)
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ Export JSON réussi !\n\nFichier : {filename}\nActivités exportées : {len(self.filtered_activities)}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export JSON :\n\n{str(e)}"
            )
    
    # ==========================================
    # STYLES
    # ==========================================
    
    def create_stat_box(self, label, value, color):
        """Crée une boîte de statistique."""
        box = QWidget()
        box.setStyleSheet(f"""
            QWidget {{
                background-color: {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)
        
        layout = QVBoxLayout(box)
        
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 20px; font-weight: bold; color: white;")
        value_label.setAlignment(Qt.AlignCenter)
        
        label_label = QLabel(label)
        label_label.setStyleSheet("font-size: 11px; color: white;")
        label_label.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(value_label)
        layout.addWidget(label_label)
        
        return box
    
    def get_input_style(self):
        """Style des inputs."""
        return """
            QLineEdit, QComboBox {
                padding: 12px;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                font-size: 14px;
                background: white;
            }
        """
    
    def get_button_style(self, color):
        """Style des boutons."""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """
    
    def get_table_style(self):
        """Style de la table."""
        return """
            QTableWidget {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background-color: white;
                gridline-color: #F0F0F0;
            }
            QTableWidget::item {
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 12px;
                border: none;
                font-weight: bold;
                color: #333;
            }
            QTableWidget::item:alternate {
                background-color: #FAFAFA;
            }
        """
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
