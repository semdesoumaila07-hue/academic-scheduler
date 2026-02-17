"""
Onglet d'analyse et retards - VERSION COMPLÈTE UC7
Module complet d'analyse des retards académiques
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QTableWidget, QTableWidgetItem, QHeaderView, QFrame,
    QProgressBar, QDialog, QFormLayout, QComboBox, QDateEdit,
    QMessageBox, QFileDialog, QTextEdit, QGroupBox, QCheckBox,
    QSpinBox
)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QColor, QFont
from datetime import datetime, timedelta
import json
from pathlib import Path
import random


class AnalysisDialog(QDialog):
    """Dialogue pour configurer l'analyse des retards"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Configuration de l'Analyse des Retards")
        self.setModal(True)
        self.setMinimumWidth(600)
        self.setMinimumHeight(500)
        self.init_ui()
    
    def init_ui(self):
        """Initialiser l'interface"""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # Titre
        title = QLabel("📊 Configuration de l'Analyse")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #1a1a1a;")
        layout.addWidget(title)
        
        subtitle = QLabel("Choisissez le niveau de granularité et la période d'analyse")
        subtitle.setStyleSheet("font-size: 14px; color: #666;")
        layout.addWidget(subtitle)
        
        layout.addSpacing(10)
        
        # Formulaire
        form_layout = QFormLayout()
        form_layout.setSpacing(15)
        
        # Niveau de granularité
        self.granularite_combo = QComboBox()
        self.granularite_combo.addItems([
            "Par Activité (détaillé)",
            "Par Classe/Cohorte",
            "Par Parcours",
            "Par UFR",
            "Par Université (global)"
        ])
        self.granularite_combo.setFixedHeight(40)
        self.granularite_combo.setStyleSheet("""
            QComboBox {
                padding: 10px;
                border: 2px solid #3498db;
                border-radius: 6px;
                font-size: 14px;
                background: white;
            }
        """)
        form_layout.addRow("Niveau de granularité *:", self.granularite_combo)
        
        # Période d'analyse
        periode_group = QGroupBox("📅 Période d'analyse")
        periode_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                color: #27ae60;
            }
        """)
        
        periode_layout = QFormLayout(periode_group)
        
        self.date_debut = QDateEdit()
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDate(QDate.currentDate().addMonths(-3))
        self.date_debut.setFixedHeight(40)
        self.date_debut.setDisplayFormat("dd/MM/yyyy")
        periode_layout.addRow("Date début:", self.date_debut)
        
        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDate(QDate.currentDate())
        self.date_fin.setFixedHeight(40)
        self.date_fin.setDisplayFormat("dd/MM/yyyy")
        periode_layout.addRow("Date fin:", self.date_fin)
        
        layout.addWidget(periode_group)
        
        # Filtres
        filtres_group = QGroupBox("🔍 Filtres (optionnel)")
        filtres_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                border: 2px solid #9b59b6;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 15px;
            }
            QGroupBox::title {
                color: #9b59b6;
            }
        """)
        
        filtres_layout = QFormLayout(filtres_group)
        
        # Filtre type d'activité
        self.type_combo = QComboBox()
        self.type_combo.addItems(["Tous les types", "CM uniquement", "TD uniquement", "TP uniquement"])
        self.type_combo.setFixedHeight(40)
        filtres_layout.addRow("Type d'activité:", self.type_combo)
        
        # Seuil d'urgence
        self.seuil_spin = QSpinBox()
        self.seuil_spin.setRange(0, 100)
        self.seuil_spin.setValue(50)
        self.seuil_spin.setSuffix(" %")
        self.seuil_spin.setFixedHeight(40)
        filtres_layout.addRow("Seuil urgence (α):", self.seuil_spin)
        
        layout.addWidget(filtres_group)
        
        layout.addSpacing(10)
        
        # Boutons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        btn_cancel = QPushButton("❌ Annuler")
        btn_cancel.setFixedSize(140, 45)
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #e0e0e0;
                color: #333;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d0d0d0;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        btn_analyze = QPushButton("📊 Analyser")
        btn_analyze.setFixedSize(180, 45)
        btn_analyze.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        btn_analyze.clicked.connect(self.accept)
        
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_analyze)
        
        layout.addLayout(btn_layout)
    
    def get_config(self):
        """Récupérer la configuration"""
        return {
            'granularite': self.granularite_combo.currentText(),
            'date_debut': self.date_debut.date().toString("dd/MM/yyyy"),
            'date_fin': self.date_fin.date().toString("dd/MM/yyyy"),
            'type_filtre': self.type_combo.currentText(),
            'seuil_urgence': self.seuil_spin.value()
        }


class AnalysisTab(QWidget):
    """Onglet pour l'analyse des retards - VERSION COMPLÈTE UC7"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # Données
        self.current_analysis = None
        self.retards_data = []
        
        self.init_ui()
        self.load_sample_data()
    
    def init_ui(self):
        """Initialise l'interface."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
        
        # ==========================================
        # EN-TÊTE
        # ==========================================
        header_layout = QVBoxLayout()
        
        title = QLabel("Analyse des Retards Académiques")
        title.setStyleSheet("font-size: 32px; font-weight: bold; color: #1a1a1a;")
        header_layout.addWidget(title)
        
        subtitle = QLabel("UC7 - Suivi de la progression et détection des activités en retard")
        subtitle.setStyleSheet("font-size: 16px; color: #666;")
        header_layout.addWidget(subtitle)
        
        layout.addLayout(header_layout)
        
        # ==========================================
        # CARTES DE STATISTIQUES
        # ==========================================
        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(20)
        
        self.stat_urgent = self.create_stat_card("Activités urgentes", "0", "#EF4444", "α ≥ 0.5")
        self.stat_critical = self.create_stat_card("Activités critiques", "0", "#F59E0B", "α ≥ 1.0")
        self.stat_retard_global = self.create_stat_card("Retard global", "0h", "#6366F1", "Total")
        self.stat_progression = self.create_stat_card("Taux progression", "0%", "#10B981", "Moyen")
        
        self.stats_layout.addWidget(self.stat_urgent)
        self.stats_layout.addWidget(self.stat_critical)
        self.stats_layout.addWidget(self.stat_retard_global)
        self.stats_layout.addWidget(self.stat_progression)
        
        layout.addLayout(self.stats_layout)
        
        # ==========================================
        # BOUTONS D'ACTION
        # ==========================================
        btn_layout = QHBoxLayout()
        
        self.btn_configure = QPushButton("⚙️ Configurer l'analyse")
        self.btn_configure.setStyleSheet(self.get_button_style("#3498db"))
        self.btn_configure.setFixedHeight(45)
        self.btn_configure.clicked.connect(self.configure_analysis)  # ✅ CONNECTÉ
        
        self.btn_calculate = QPushButton("📊 Calculer les retards")
        self.btn_calculate.setStyleSheet(self.get_button_style("#27ae60"))
        self.btn_calculate.setFixedHeight(45)
        self.btn_calculate.clicked.connect(self.calculate_delays)  # ✅ CONNECTÉ
        
        self.btn_stats = QPushButton("📈 Statistiques détaillées")
        self.btn_stats.setStyleSheet(self.get_button_style("#9b59b6"))
        self.btn_stats.setFixedHeight(45)
        self.btn_stats.clicked.connect(self.show_detailed_stats)  # ✅ CONNECTÉ
        
        self.btn_export = QPushButton("📥 Exporter rapport")
        self.btn_export.setStyleSheet(self.get_button_style("#e67e22"))
        self.btn_export.setFixedHeight(45)
        self.btn_export.clicked.connect(self.export_report)  # ✅ CONNECTÉ
        
        btn_layout.addWidget(self.btn_configure)
        btn_layout.addWidget(self.btn_calculate)
        btn_layout.addWidget(self.btn_stats)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addStretch()
        
        layout.addLayout(btn_layout)
        
        # ==========================================
        # FILTRES RAPIDES
        # ==========================================
        filter_layout = QHBoxLayout()
        
        filter_label = QLabel("Filtres rapides:")
        filter_label.setStyleSheet("font-weight: bold;")
        filter_layout.addWidget(filter_label)
        
        self.filter_type = QComboBox()
        self.filter_type.addItems(["Tous les types", "CM", "TD", "TP"])
        self.filter_type.setFixedHeight(35)
        self.filter_type.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_type)
        
        self.filter_urgence = QComboBox()
        self.filter_urgence.addItems(["Toutes urgences", "Urgentes uniquement (α≥0.5)", "Critiques uniquement (α≥1.0)"])
        self.filter_urgence.setFixedHeight(35)
        self.filter_urgence.currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(self.filter_urgence)
        
        filter_layout.addStretch()
        
        layout.addLayout(filter_layout)
        
        # ==========================================
        # TABLE DES RETARDS
        # ==========================================
        table_header = QHBoxLayout()
        
        table_label = QLabel("📋 Activités avec retards détectés")
        table_label.setStyleSheet("font-size: 18px; font-weight: bold;")
        table_header.addWidget(table_label)
        
        table_header.addStretch()
        
        self.count_label = QLabel("0 résultat(s)")
        self.count_label.setStyleSheet("color: #666; font-size: 14px;")
        table_header.addWidget(self.count_label)
        
        layout.addLayout(table_header)
        
        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "Activité", "Type", "Cohorte", "Volume (h)", "Réalisé (h)",
            "Retard (h)", "Lag (j)", "Urgence (α)", "Actions"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(8, QHeaderView.Fixed)
        self.table.setColumnWidth(8, 100)
        self.table.setStyleSheet(self.get_modern_table_style())
        self.table.setAlternatingRowColors(True)
        
        layout.addWidget(self.table)
    
    def load_sample_data(self):
        """Charger des données exemple"""
        self.retards_data = [
            {
                'activite': 'Développement Web',
                'type': 'CM',
                'cohorte': 'L3 Info',
                'volume_heures': 25,
                'heures_realisees': 10,
                'retard_heures': 15,
                'lag_jours': 3.5,
                'alpha': 0.65,
                'progression': 40
            },
            {
                'activite': 'Bases de données',
                'type': 'TP',
                'cohorte': 'L3 Info',
                'volume_heures': 20,
                'heures_realisees': 12,
                'retard_heures': 8,
                'lag_jours': 2.0,
                'alpha': 0.52,
                'progression': 60
            },
            {
                'activite': 'Réseaux informatiques',
                'type': 'TD',
                'cohorte': 'L2 Info',
                'volume_heures': 20,
                'heures_realisees': 18,
                'retard_heures': 2,
                'lag_jours': 0.5,
                'alpha': 0.25,
                'progression': 90
            },
            {
                'activite': 'Intelligence Artificielle',
                'type': 'CM',
                'cohorte': 'M1 Info',
                'volume_heures': 30,
                'heures_realisees': 5,
                'retard_heures': 25,
                'lag_jours': 5.0,
                'alpha': 1.2,
                'progression': 17
            },
        ]
        
        self.apply_filters()
        self.update_statistics()
    
    def apply_filters(self):
        """Appliquer les filtres"""
        type_filter = self.filter_type.currentText()
        urgence_filter = self.filter_urgence.currentText()
        
        filtered_data = []
        
        for item in self.retards_data:
            # Filtre de type
            if type_filter != "Tous les types" and item['type'] != type_filter:
                continue
            
            # Filtre d'urgence
            if urgence_filter == "Urgentes uniquement (α≥0.5)" and item['alpha'] < 0.5:
                continue
            elif urgence_filter == "Critiques uniquement (α≥1.0)" and item['alpha'] < 1.0:
                continue
            
            filtered_data.append(item)
        
        self.refresh_table(filtered_data)
        self.count_label.setText(f"{len(filtered_data)} résultat(s)")
    
    def refresh_table(self, data):
        """Rafraîchir la table"""
        self.table.setRowCount(len(data))
        
        for i, item in enumerate(data):
            # Activité
            self.table.setItem(i, 0, QTableWidgetItem(item['activite']))
            
            # Type
            type_item = QTableWidgetItem(item['type'])
            if item['type'] == "CM":
                type_item.setBackground(QColor("#E3F2FD"))
            elif item['type'] == "TD":
                type_item.setBackground(QColor("#E8F5E9"))
            elif item['type'] == "TP":
                type_item.setBackground(QColor("#FFF3E0"))
            self.table.setItem(i, 1, type_item)
            
            # Cohorte
            self.table.setItem(i, 2, QTableWidgetItem(item['cohorte']))
            
            # Volume
            self.table.setItem(i, 3, QTableWidgetItem(f"{item['volume_heures']}h"))
            
            # Réalisé
            self.table.setItem(i, 4, QTableWidgetItem(f"{item['heures_realisees']}h"))
            
            # Retard
            retard_item = QTableWidgetItem(f"{item['retard_heures']}h")
            if item['retard_heures'] > 10:
                retard_item.setBackground(QColor("#FFEBEE"))
            self.table.setItem(i, 5, retard_item)
            
            # Lag
            self.table.setItem(i, 6, QTableWidgetItem(f"{item['lag_jours']:.1f}j"))
            
            # Alpha (urgence)
            alpha_item = QTableWidgetItem(f"{item['alpha']:.2f}")
            if item['alpha'] >= 1.0:
                alpha_item.setBackground(QColor("#FFCDD2"))
                alpha_item.setForeground(QColor("#C62828"))
            elif item['alpha'] >= 0.5:
                alpha_item.setBackground(QColor("#FFF3C7"))
                alpha_item.setForeground(QColor("#F57C00"))
            else:
                alpha_item.setBackground(QColor("#E8F5E9"))
                alpha_item.setForeground(QColor("#2E7D32"))
            self.table.setItem(i, 7, alpha_item)
            
            # Actions
            actions_widget = self.create_action_buttons(item)
            self.table.setCellWidget(i, 8, actions_widget)
    
    def update_statistics(self):
        """Mettre à jour les statistiques"""
        if not self.retards_data:
            return
        
        # Compter urgentes et critiques
        urgentes = len([r for r in self.retards_data if r['alpha'] >= 0.5])
        critiques = len([r for r in self.retards_data if r['alpha'] >= 1.0])
        
        # Retard global
        retard_total = sum(r['retard_heures'] for r in self.retards_data)
        
        # Progression moyenne
        prog_moyenne = sum(r['progression'] for r in self.retards_data) / len(self.retards_data)
        
        # Mettre à jour les cartes
        self.update_stat_card(self.stat_urgent, str(urgentes))
        self.update_stat_card(self.stat_critical, str(critiques))
        self.update_stat_card(self.stat_retard_global, f"{retard_total}h")
        self.update_stat_card(self.stat_progression, f"{prog_moyenne:.0f}%")
    
    def update_stat_card(self, card, value):
        """Mettre à jour une carte de statistique"""
        labels = card.findChildren(QLabel)
        if labels:
            labels[0].setText(value)
    
    # ==========================================
    # ACTIONS
    # ==========================================
    
    def configure_analysis(self):
        """Configurer l'analyse"""
        dialog = AnalysisDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            config = dialog.get_config()
            
            QMessageBox.information(
                self,
                "Configuration enregistrée",
                f"✅ Analyse configurée :\n\n"
                f"Granularité : {config['granularite']}\n"
                f"Période : {config['date_debut']} → {config['date_fin']}\n"
                f"Filtre : {config['type_filtre']}\n"
                f"Seuil urgence : {config['seuil_urgence']}%"
            )
            
            self.current_analysis = config
    
    def calculate_delays(self):
        """Calculer les retards"""
        if not self.retards_data:
            QMessageBox.warning(self, "Attention", "Aucune donnée disponible pour le calcul !")
            return
        
        # Calculer les statistiques
        total_activites = len(self.retards_data)
        retard_total = sum(r['retard_heures'] for r in self.retards_data)
        retard_moyen = retard_total / total_activites if total_activites > 0 else 0
        retard_max = max(r['retard_heures'] for r in self.retards_data)
        
        # Écart-type
        moyenne = retard_moyen
        variance = sum((r['retard_heures'] - moyenne) ** 2 for r in self.retards_data) / total_activites
        ecart_type = variance ** 0.5
        
        # Classement
        classement = sorted(self.retards_data, key=lambda x: x['retard_heures'], reverse=True)[:5]
        
        # Message
        message = f"""
📊 CALCUL DES RETARDS ACADÉMIQUES

📈 STATISTIQUES GLOBALES :
- Nombre d'activités analysées : {total_activites}
- Retard total : {retard_total}h
- Retard moyen : {retard_moyen:.2f}h
- Retard maximum : {retard_max}h
- Écart-type : {ecart_type:.2f}h

🏆 TOP 5 DES ACTIVITÉS EN RETARD :
"""
        
        for i, act in enumerate(classement, 1):
            message += f"\n{i}. {act['activite']} ({act['cohorte']}) : {act['retard_heures']}h (α={act['alpha']:.2f})"
        
        QMessageBox.information(self, "Résultats du Calcul", message)
    
    def show_detailed_stats(self):
        """Afficher les statistiques détaillées"""
        if not self.retards_data:
            QMessageBox.warning(self, "Attention", "Aucune donnée disponible !")
            return
        
        # Préparer les statistiques
        total = len(self.retards_data)
        cm = len([r for r in self.retards_data if r['type'] == 'CM'])
        td = len([r for r in self.retards_data if r['type'] == 'TD'])
        tp = len([r for r in self.retards_data if r['type'] == 'TP'])
        
        retard_total = sum(r['retard_heures'] for r in self.retards_data)
        retard_moyen = retard_total / total
        retard_max = max(r['retard_heures'] for r in self.retards_data)
        retard_min = min(r['retard_heures'] for r in self.retards_data)
        
        # Retard par type
        retard_cm = sum(r['retard_heures'] for r in self.retards_data if r['type'] == 'CM')
        retard_td = sum(r['retard_heures'] for r in self.retards_data if r['type'] == 'TD')
        retard_tp = sum(r['retard_heures'] for r in self.retards_data if r['type'] == 'TP')
        
        # Variance et écart-type
        moyenne = retard_moyen
        variance = sum((r['retard_heures'] - moyenne) ** 2 for r in self.retards_data) / total
        ecart_type = variance ** 0.5
        
        stats = f"""
📊 STATISTIQUES DÉTAILLÉES DES RETARDS

═══════════════════════════════════════════

📈 RÉPARTITION PAR TYPE D'ACTIVITÉ :
- CM : {cm} activités ({cm/total*100:.1f}%) - Retard total : {retard_cm}h
- TD : {td} activités ({td/total*100:.1f}%) - Retard total : {retard_td}h
- TP : {tp} activités ({tp/total*100:.1f}%) - Retard total : {retard_tp}h

═══════════════════════════════════════════

📉 INDICATEURS STATISTIQUES :
- Retard total : {retard_total}h
- Retard moyen : {retard_moyen:.2f}h
- Retard médian : {sorted([r['retard_heures'] for r in self.retards_data])[total//2]:.2f}h
- Retard minimum : {retard_min}h
- Retard maximum : {retard_max}h
- Écart-type : {ecart_type:.2f}h
- Variance : {variance:.2f}

═══════════════════════════════════════════

⚠️ NIVEAUX D'URGENCE (α) :
- Critiques (α ≥ 1.0) : {len([r for r in self.retards_data if r['alpha'] >= 1.0])} activités
- Urgentes (0.5 ≤ α < 1.0) : {len([r for r in self.retards_data if 0.5 <= r['alpha'] < 1.0])} activités
- Modérées (α < 0.5) : {len([r for r in self.retards_data if r['alpha'] < 0.5])} activités

═══════════════════════════════════════════

📊 PROGRESSION GLOBALE :
- Volume horaire total : {sum(r['volume_heures'] for r in self.retards_data)}h
- Heures réalisées : {sum(r['heures_realisees'] for r in self.retards_data)}h
- Taux de progression : {sum(r['progression'] for r in self.retards_data)/total:.1f}%
        """
        
        # Dialogue pour afficher
        dialog = QDialog(self)
        dialog.setWindowTitle("Statistiques Détaillées")
        dialog.setMinimumWidth(600)
        dialog.setMinimumHeight(500)
        
        layout = QVBoxLayout(dialog)
        
        text = QTextEdit()
        text.setPlainText(stats)
        text.setReadOnly(True)
        text.setStyleSheet("font-family: 'Courier New'; font-size: 12px;")
        layout.addWidget(text)
        
        btn_close = QPushButton("Fermer")
        btn_close.clicked.connect(dialog.close)
        layout.addWidget(btn_close)
        
        dialog.exec_()
    
    def export_report(self):
        """Exporter le rapport"""
        if not self.retards_data:
            QMessageBox.warning(self, "Attention", "Aucune donnée à exporter !")
            return
        
        reply = QMessageBox.question(
            self,
            "Format d'export",
            "Choisissez le format d'export :\n\n"
            "Oui = Excel (.xlsx)\n"
            "Non = CSV (.csv)",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            self.export_to_excel()
        else:
            self.export_to_csv()
    
    def export_to_excel(self):
        """Exporter en Excel"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter en Excel",
            f"analyse_retards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not filename:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Analyse des Retards"
            
            # En-têtes
            headers = ["Activité", "Type", "Cohorte", "Volume (h)", "Réalisé (h)",
                      "Retard (h)", "Lag (j)", "Urgence (α)", "Progression (%)"]
            ws.append(headers)
            
            # Style des en-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            for item in self.retards_data:
                ws.append([
                    item['activite'],
                    item['type'],
                    item['cohorte'],
                    item['volume_heures'],
                    item['heures_realisees'],
                    item['retard_heures'],
                    round(item['lag_jours'], 2),
                    round(item['alpha'], 2),
                    item['progression']
                ])
            
            # Ajuster colonnes
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ Export Excel réussi !\n\nFichier : {filename}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self,
                "Module manquant",
                "Le module 'openpyxl' n'est pas installé.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export :\n\n{str(e)}"
            )
    
    def export_to_csv(self):
        """Exporter en CSV"""
        filename, _ = QFileDialog.getSaveFileName(
            self,
            "Exporter en CSV",
            f"analyse_retards_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV Files (*.csv)"
        )
        
        if not filename:
            return
        
        try:
            import csv
            
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                
                # En-têtes
                writer.writerow(["Activité", "Type", "Cohorte", "Volume (h)", "Réalisé (h)",
                               "Retard (h)", "Lag (j)", "Urgence (α)", "Progression (%)"])
                
                # Données
                for item in self.retards_data:
                    writer.writerow([
                        item['activite'],
                        item['type'],
                        item['cohorte'],
                        item['volume_heures'],
                        item['heures_realisees'],
                        item['retard_heures'],
                        round(item['lag_jours'], 2),
                        round(item['alpha'], 2),
                        item['progression']
                    ])
            
            QMessageBox.information(
                self,
                "Succès",
                f"✅ Export CSV réussi !\n\nFichier : {filename}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export :\n\n{str(e)}"
            )
    
    # ==========================================
    # UTILITAIRES
    # ==========================================
    
    def create_stat_card(self, title, value, color, subtitle):
        """Crée une carte de statistique."""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background: white;
                border: 1px solid #E5E7EB;
                border-left: 4px solid {color};
                border-radius: 8px;
                padding: 20px;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(8)
        
        value_label = QLabel(value)
        value_label.setStyleSheet(f"font-size: 28px; font-weight: 700; color: {color};")
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 14px; color: #6B7280; font-weight: 500;")
        
        subtitle_label = QLabel(subtitle)
        subtitle_label.setStyleSheet("font-size: 12px; color: #9CA3AF;")
        
        layout.addWidget(value_label)
        layout.addWidget(title_label)
        layout.addWidget(subtitle_label)
        
        return card
    
    def create_action_buttons(self, item):
        """Créer les boutons d'action"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(5)
        
        btn_view = QPushButton("👁️")
        btn_view.setFixedSize(30, 30)
        btn_view.setStyleSheet("""
            QPushButton {
                background: transparent;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #E3F2FD;
                border-radius: 4px;
            }
        """)
        btn_view.setCursor(Qt.PointingHandCursor)
        btn_view.clicked.connect(lambda: self.view_details(item))
        
        layout.addWidget(btn_view)
        layout.addStretch()
        
        return widget
    
    def view_details(self, item):
        """Voir les détails d'une activité"""
        details = f"""
📋 DÉTAILS DE L'ACTIVITÉ

Activité : {item['activite']}
Type : {item['type']}
Cohorte : {item['cohorte']}

📊 VOLUMES :
- Volume horaire prévu : {item['volume_heures']}h
- Heures réalisées : {item['heures_realisees']}h
- Retard : {item['retard_heures']}h

⏱️ INDICATEURS :
- Lag (jours) : {item['lag_jours']:.1f} jours
- Urgence (α) : {item['alpha']:.2f}
- Progression : {item['progression']}%

⚠️ ÉTAT :
"""
        if item['alpha'] >= 1.0:
            details += "• 🔴 CRITIQUE - Action immédiate requise !"
        elif item['alpha'] >= 0.5:
            details += "• 🟡 URGENTE - Attention nécessaire"
        else:
            details += "• 🟢 MODÉRÉE - Suivi régulier"
        
        QMessageBox.information(self, "Détails de l'activité", details)
    
    def get_button_style(self, color):
        """Style des boutons"""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                opacity: 0.9;
            }}
        """
    
    def get_modern_table_style(self):
        """Style moderne de la table."""
        return """
            QTableWidget {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background-color: white;
                gridline-color: #F0F0F0;
            }
            QTableWidget::item {
                padding: 12px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 12px;
                border: none;
                font-weight: bold;
                font-size: 13px;
                color: #333;
            }
            QTableWidget::item:alternate {
                background-color: #FAFAFA;
            }
        """