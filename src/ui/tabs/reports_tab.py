# -*- coding: utf-8 -*-
"""
Onglet Rapports — UC8 — Graphiques + PDF + Excel + Export automatique vendredi
"""
import os, io, time
from datetime import datetime, date
from pathlib import Path

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QFrame, QComboBox, QDateEdit, QProgressBar,
    QMessageBox, QFileDialog, QScrollArea, QTabWidget
)
from PyQt5.QtCore import Qt, QDate, QThread, pyqtSignal, QTimer
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from src.database.db_manager import db_manager
from src.database.models import (
    AcademicActivityModel, ScheduleSlotModel, TeacherModel,
    CohortModel, LeaveRequestModel
)

COLORS = ["#3B82F6","#10B981","#F59E0B","#EF4444","#8B5CF6","#06B6D4","#84CC16","#F97316"]

def _fig(w=7, h=3.5):
    fig = Figure(figsize=(w,h), facecolor="white")
    fig.subplots_adjust(left=0.12, right=0.95, top=0.88, bottom=0.22)
    return fig

def chart_charge(session):
    teachers = session.query(TeacherModel).all()
    acts_all = session.query(AcademicActivityModel).all()
    noms, vols, dones = [], [], []
    for t in teachers:
        acts = [a for a in acts_all if getattr(a,"teacher_id",None)==t.id]
        v = sum(a.volume_hours or 0 for a in acts)
        d = sum(a.hours_done   or 0 for a in acts)
        if v > 0:
            noms.append(t.full_name.split()[-1][:10])
            vols.append(v); dones.append(d)
    fig = _fig(); ax = fig.add_subplot(111)
    x = range(len(noms)); w = 0.35
    ax.bar([i-w/2 for i in x], vols,  w, label="Volume",     color="#3B82F6", alpha=0.85)
    ax.bar([i+w/2 for i in x], dones, w, label="Planifiees", color="#10B981", alpha=0.85)
    ax.set_xticks(list(x)); ax.set_xticklabels(noms, rotation=30, ha="right", fontsize=8)
    ax.set_title("Charge par enseignant (h)", fontsize=11, fontweight="bold")
    ax.legend(fontsize=8); ax.set_ylabel("Heures"); ax.grid(axis="y", alpha=0.3)
    return fig

def chart_repartition(session):
    acts = session.query(AcademicActivityModel).all()
    types = {}
    for a in acts:
        t = getattr(a,"type",None)
        lbl = t.value if hasattr(t,"value") else str(t or "Autre")
        types[lbl] = types.get(lbl,0) + (a.volume_hours or 0)
    fig = _fig(6,3.5); ax = fig.add_subplot(111)
    if types:
        ax.pie(types.values(), labels=types.keys(), colors=COLORS[:len(types)],
               autopct="%1.0f%%", startangle=90, textprops={"fontsize":8})
    ax.set_title("Repartition par type (heures)", fontsize=11, fontweight="bold")
    return fig

def chart_hebdo(session):
    slots = session.query(ScheduleSlotModel).all()
    sem = {}
    for s in slots:
        if s.date:
            try:
                d = s.date if isinstance(s.date, date) else datetime.strptime(str(s.date),"%Y-%m-%d").date()
                iso = d.isocalendar()
                k = f"S{iso[1]:02d}\n{iso[0]}"
                sem[k] = sem.get(k,0)+1
            except: pass
    sem_s = dict(sorted(sem.items()))
    keys = list(sem_s.keys())[-12:]; vals = [sem_s[k] for k in keys]
    fig = _fig(); ax = fig.add_subplot(111)
    ax.plot(range(len(keys)), vals, color="#8B5CF6", linewidth=2, marker="o", markersize=5)
    ax.fill_between(range(len(keys)), vals, alpha=0.15, color="#8B5CF6")
    ax.set_xticks(range(len(keys))); ax.set_xticklabels(keys, fontsize=7)
    ax.set_title("Evolution hebdomadaire des creneaux", fontsize=11, fontweight="bold")
    ax.set_ylabel("Creneaux"); ax.grid(alpha=0.3)
    return fig

def chart_conges(session):
    try:
        leaves = session.query(LeaveRequestModel).all()
        def _s(l):
            return str(getattr(l.status,"value",l.status) or "").strip()
        app = sum(1 for l in leaves if _s(l) == "Approuvé")
        ref = sum(1 for l in leaves if _s(l) == "Rejeté")
        ann = sum(1 for l in leaves if _s(l) == "Annulé")
        att = max(0, len(leaves) - app - ref - ann)
        fig = _fig(6,3); ax = fig.add_subplot(111)
        bars = ax.barh(["Approuvees","Refusees","Annulees","En attente"], [app,ref,ann,att],
                       color=["#10B981","#EF4444","#9CA3AF","#F59E0B"], alpha=0.85)
        for bar, val in zip(bars,[app,ref,ann,att]):
            ax.text(bar.get_width()+0.1, bar.get_y()+bar.get_height()/2, str(val), va="center", fontsize=9)
        ax.set_title("Statuts des conges", fontsize=11, fontweight="bold")
        ax.grid(axis="x", alpha=0.3)
    except:
        fig = _fig(); ax = fig.add_subplot(111)
        ax.text(0.5,0.5,"Donnees non disponibles", ha="center", va="center")
    return fig

def chart_cohorte(session):
    cohortes = session.query(CohortModel).all()
    noms, counts = [], []
    for c in cohortes:
        n = session.query(ScheduleSlotModel).filter_by(cohort_id=c.id).count()
        if n > 0:
            noms.append(getattr(c,"name","C"+str(c.id))[:12])
            counts.append(n)
    fig = _fig(); ax = fig.add_subplot(111)
    if noms:
        ax.bar(range(len(noms)), counts, color=COLORS[:len(noms)], alpha=0.85)
        ax.set_xticks(range(len(noms))); ax.set_xticklabels(noms, rotation=30, ha="right", fontsize=8)
    else:
        ax.text(0.5,0.5,"Aucune donnee", ha="center", va="center")
    ax.set_title("Creneaux par cohorte", fontsize=11, fontweight="bold")
    ax.set_ylabel("Creneaux"); ax.grid(axis="y", alpha=0.3)
    return fig

def do_export_pdf(filename, session, p_start, p_end, rtype):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    doc = SimpleDocTemplate(filename, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)
    styles = getSampleStyleSheet()
    ts = ParagraphStyle("T", parent=styles["Title"], fontSize=18, textColor=colors.HexColor("#1F4E79"), alignment=TA_CENTER)
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=12, textColor=colors.HexColor("#1F4E79"))
    story = []
    story.append(Paragraph(f"Rapport - {rtype}", ts))
    story.append(Paragraph(f"Periode : {p_start} - {p_end}", styles["Normal"]))
    story.append(Paragraph(f"Genere le : {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))
    acts = session.query(AcademicActivityModel).all()
    slots = session.query(ScheduleSlotModel).all()
    teachers = session.query(TeacherModel).all()
    cohortes = session.query(CohortModel).all()
    vol  = sum(a.volume_hours or 0 for a in acts)
    done = sum(a.hours_done   or 0 for a in acts)
    taux = round(done/vol*100,1) if vol else 0
    story.append(Paragraph("Indicateurs cles", h1))
    data = [["Indicateur","Valeur"],
            ["Activites", str(len(acts))], ["Creneaux", str(len(slots))],
            ["Enseignants", str(len(teachers))], ["Cohortes", str(len(cohortes))],
            ["Heures planifiees", f"{done:.0f}h / {vol:.0f}h"],
            ["Taux couverture", f"{taux}%"]]
    t = Table(data, colWidths=[10*cm,6*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1F4E79")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
        ("FONTSIZE",(0,0),(-1,-1),10),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#EFF6FF")]),
        ("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D1D5DB")),
        ("PADDING",(0,0),(-1,-1),6),
    ]))
    story.append(t); story.append(Spacer(1,0.5*cm))
    for title, func in [("Charge par enseignant",chart_charge),
                         ("Repartition activites",chart_repartition),
                         ("Evolution hebdomadaire",chart_hebdo),
                         ("Statuts conges",chart_conges),
                         ("Creneaux par cohorte",chart_cohorte)]:
        story.append(Paragraph(title, h1))
        try:
            fig = func(session)
            buf = io.BytesIO()
            fig.savefig(buf, format="png", dpi=100, bbox_inches="tight")
            buf.seek(0)
            story.append(Image(buf, width=15*cm, height=7.5*cm))
            plt.close(fig)
        except Exception as e:
            story.append(Paragraph(f"Graphique non disponible: {e}", styles["Normal"]))
        story.append(Spacer(1,0.3*cm))
    doc.build(story)

def do_export_excel(filename, session, p_start, p_end):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    wb = Workbook()
    hf = Font(bold=True, color="FFFFFF")
    hb = PatternFill("solid", fgColor="1F4E79")
    alt = PatternFill("solid", fgColor="EFF6FF")
    ws = wb.active; ws.title = "Resume"
    ws["A1"] = "Rapport Academique"; ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws["A2"] = f"Periode: {p_start} - {p_end}"
    ws["A3"] = f"Genere: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws.append([])
    acts = session.query(AcademicActivityModel).all()
    slots = session.query(ScheduleSlotModel).all()
    teachers = session.query(TeacherModel).all()
    cohortes = session.query(CohortModel).all()
    vol  = sum(a.volume_hours or 0 for a in acts)
    done = sum(a.hours_done   or 0 for a in acts)
    ws.append(["Indicateur","Valeur"])
    for c in ws[ws.max_row]: c.font=hf; c.fill=hb
    rows = [("Activites",len(acts)),("Creneaux",len(slots)),
            ("Enseignants",len(teachers)),("Cohortes",len(cohortes)),
            ("Heures planifiees",round(done,1)),("Volume total",round(vol,1)),
            ("Taux couverture %",round(done/vol*100,1) if vol else 0)]
    for i,(k,v) in enumerate(rows):
        ws.append([k,v])
        if i%2==0:
            for c in ws[ws.max_row]: c.fill=alt
    ws.column_dimensions["A"].width=28; ws.column_dimensions["B"].width=18
    ws2 = wb.create_sheet("Charge Enseignants")
    ws2.append(["Enseignant","Volume (h)","Planifiees (h)","Taux (%)"])
    for c in ws2[1]: c.font=hf; c.fill=hb
    for t in teachers:
        t_acts=[a for a in acts if getattr(a,"teacher_id",None)==t.id]
        v=sum(a.volume_hours or 0 for a in t_acts); d=sum(a.hours_done or 0 for a in t_acts)
        ws2.append([t.full_name, round(v,1), round(d,1), round(d/v*100,1) if v else 0])
    for col in ["A","B","C","D"]: ws2.column_dimensions[col].width=22
    ws3 = wb.create_sheet("Activites")
    ws3.append(["Nom","Type","Volume (h)","Planifiees (h)","Alpha"])
    for c in ws3[1]: c.font=hf; c.fill=hb
    for a in acts:
        v=a.volume_hours or 0; d=a.hours_done or 0
        tp = str(getattr(getattr(a,"type",None),"value",None) or "N/A")
        ws3.append([a.name, tp, round(v,1), round(d,1), round((v-d)/v,2) if v else 0])
    for col in ["A","B","C","D","E"]: ws3.column_dimensions[col].width=22
    wb.save(filename)


class ReportThread(QThread):
    progress = pyqtSignal(int, str)
    done     = pyqtSignal(dict)
    error    = pyqtSignal(str)
    def __init__(self, fmt, fname, session, ps, pe, rtype):
        super().__init__()
        self.fmt=fmt; self.fname=fname; self.session=session
        self.ps=ps; self.pe=pe; self.rtype=rtype
    def run(self):
        try:
            self.progress.emit(20,"Collecte des donnees...")
            time.sleep(0.2)
            self.progress.emit(50,f"Generation {self.fmt}...")
            Path(self.fname).parent.mkdir(parents=True, exist_ok=True)
            if self.fmt=="PDF":
                do_export_pdf(self.fname, self.session, self.ps, self.pe, self.rtype)
            elif self.fmt=="Excel":
                do_export_excel(self.fname, self.session, self.ps, self.pe)
            else:
                base = self.fname.rsplit(".",1)[0]
                do_export_pdf(base+".pdf", self.session, self.ps, self.pe, self.rtype)
                self.progress.emit(75,"Export Excel...")
                do_export_excel(base+".xlsx", self.session, self.ps, self.pe)
            self.progress.emit(100,"Rapport genere!")
            self.done.emit({"filename":self.fname})
        except Exception as e:
            import traceback; self.error.emit(str(e)+"\n"+traceback.format_exc())


class ReportsTab(QWidget):
    def __init__(self, current_user=None, parent=None):
        super().__init__(parent)
        self.current_user = current_user
        self.session      = db_manager.get_session()
        self.thread       = None
        self._build()
        self._setup_auto()

    def _setup_auto(self):
        """
        Configure le déclencheur d'export automatique.

        Le timer vérifie toutes les 30 secondes si on est vendredi entre 18h00 et 18h01.
        La fenêtre de 1 minute garantit que le déclencheur n'est pas raté même si
        le timer s'exécute légèrement en retard. Un verrou (_auto_done_today) empêche
        l'export de se déclencher plusieurs fois dans la même fenêtre.
        """
        self._auto_done_today = None   # mémorise la date du dernier export automatique
        self._atimer = QTimer(self)
        self._atimer.setInterval(30000)   # toutes les 30 secondes
        self._atimer.timeout.connect(self._check_auto)
        self._atimer.start()

    def _check_auto(self):
        """
        Vérifie si l'export automatique doit être déclenché.

        Conditions : vendredi (weekday==4) + heure 18h00–18h00 + pas encore exporté aujourd'hui.
        La fenêtre de 60 secondes (minute==0) couvre le délai du timer de 30 s.
        """
        now = datetime.now()
        today = now.date()
        is_friday_18h = (now.weekday() == 4 and now.hour == 18 and now.minute == 0)

        # Evite un double déclenchement dans la même fenêtre d'une minute
        if is_friday_18h and self._auto_done_today != today:
            self._auto_done_today = today
            self._auto_export(silent=True)

    def _auto_export(self, silent=False):
        """
        Exporte le rapport au format PDF + Excel dans le dossier Rapports_Academiques.

        Args:
            silent: Si True, n'affiche pas de boîte de dialogue de confirmation.
                    Utilisé pour l'export automatique vendredi 18h.
        """
        try:
            # Crée le dossier de destination (Documents, sinon Bureau)
            folder = os.path.join(os.path.expanduser("~"), "Documents", "Rapports_Academiques")
            try:
                os.makedirs(folder, exist_ok=True)
            except Exception:
                folder = os.path.join(os.path.expanduser("~"), "Desktop", "Rapports_Academiques")
                os.makedirs(folder, exist_ok=True)

            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            sd = self.dt_start.date()
            ed = self.dt_end.date()
            ps = f"{sd.day():02d}/{sd.month():02d}/{sd.year()}"
            pe = f"{ed.day():02d}/{ed.month():02d}/{ed.year()}"

            pdf_path  = os.path.join(folder, f"rapport_{stamp}.pdf")
            xlsx_path = os.path.join(folder, f"rapport_{stamp}.xlsx")

            do_export_pdf(pdf_path,  self.session, ps, pe, "Export auto vendredi")
            do_export_excel(xlsx_path, self.session, ps, pe)

            if not silent:
                QMessageBox.information(
                    self, "Export automatique",
                    f"Rapport vendredi 18h généré avec succès :\n{folder}"
                )
        except Exception as e:
            if not silent:
                QMessageBox.warning(self, "Erreur export auto", str(e))
            else:
                print(f"[reports_tab] export auto error: {e}")

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(16)
        title = QLabel("Rapports & Statistiques")
        title.setStyleSheet("font-size:24px; font-weight:700; color:#1F2937;")
        lay.addWidget(title)
        sub = QLabel("UC8 - Graphiques + PDF + Excel + Export automatique vendredi 18h")
        sub.setStyleSheet("font-size:12px; color:#6B7280;")
        lay.addWidget(sub)
        tabs = QTabWidget()
        tabs.setStyleSheet("QTabBar::tab{padding:7px 18px;} QTabBar::tab:selected{font-weight:bold; color:#1F4E79;}")
        t1 = self._build_charts_tab()
        t2 = self._build_export_tab()
        tabs.addTab(t1,"Graphiques")
        tabs.addTab(t2,"Generer rapport")
        lay.addWidget(tabs)

    def _build_charts_tab(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setContentsMargins(12,12,12,12)
        btn = QPushButton("Rafraichir les graphiques")
        btn.setStyleSheet("background:#1F4E79; color:white; border-radius:6px; padding:6px 16px;")
        btn.clicked.connect(self._refresh)
        hl = QHBoxLayout(); hl.addStretch(); hl.addWidget(btn); lay.addLayout(hl)
        sc = QScrollArea(); sc.setWidgetResizable(True); sc.setFrameShape(QFrame.NoFrame)
        self._cw = QWidget(); self._cl = QVBoxLayout(self._cw); self._cl.setSpacing(16)
        sc.setWidget(self._cw); lay.addWidget(sc)
        self._refresh()
        return w

    def _refresh(self):
        while self._cl.count():
            item = self._cl.takeAt(0)
            if item.widget(): item.widget().deleteLater()
        charts = [("Charge par enseignant",chart_charge),
                  ("Repartition des activites",chart_repartition),
                  ("Evolution hebdomadaire",chart_hebdo),
                  ("Statuts des conges",chart_conges),
                  ("Creneaux par cohorte",chart_cohorte)]
        for title, func in charts:
            fr = QFrame()
            fr.setStyleSheet("QFrame{background:white; border:1px solid #E5E7EB; border-radius:8px;}")
            fl = QVBoxLayout(fr); fl.setContentsMargins(10,10,10,10)
            lbl = QLabel(title); lbl.setStyleSheet("font-size:12px; font-weight:bold; color:#1F2937;")
            fl.addWidget(lbl)
            try:
                fig = func(self.session)
                cv = FigureCanvas(fig); cv.setFixedHeight(260)
                fl.addWidget(cv); plt.close(fig)
            except Exception as e:
                fl.addWidget(QLabel(f"Non disponible: {e}"))
            self._cl.addWidget(fr)

    def _build_export_tab(self):
        w = QWidget(); lay = QVBoxLayout(w); lay.setContentsMargins(20,20,20,20); lay.setSpacing(14)
        fr = QFrame()
        fr.setStyleSheet("QFrame{background:white; border:1px solid #E5E7EB; border-radius:10px; padding:16px;}")
        fl = QVBoxLayout(fr); fl.setSpacing(14)
        r1 = QHBoxLayout(); r1.setSpacing(14)
        c1 = QVBoxLayout(); c1.addWidget(QLabel("Type de rapport"))
        self.rtype_cb = QComboBox()
        self.rtype_cb.addItems(["Synthese semestrielle","Comparaison inter-UFR","Evolution annuelle"])
        self.rtype_cb.setFixedHeight(38); self.rtype_cb.setStyleSheet(self._is()); c1.addWidget(self.rtype_cb)
        c2 = QVBoxLayout(); c2.addWidget(QLabel("Format"))
        self.fmt_cb = QComboBox()
        self.fmt_cb.addItems(["PDF + Excel","PDF","Excel"])
        self.fmt_cb.setFixedHeight(38); self.fmt_cb.setStyleSheet(self._is()); c2.addWidget(self.fmt_cb)
        r1.addLayout(c1,1); r1.addLayout(c2,1); fl.addLayout(r1)
        r2 = QHBoxLayout(); r2.setSpacing(14)
        c3 = QVBoxLayout(); c3.addWidget(QLabel("Debut"))
        self.dt_start = QDateEdit(QDate(QDate.currentDate().year(),9,1))
        self.dt_start.setCalendarPopup(True); self.dt_start.setFixedHeight(38)
        self.dt_start.setStyleSheet(self._is()); c3.addWidget(self.dt_start)
        c4 = QVBoxLayout(); c4.addWidget(QLabel("Fin"))
        self.dt_end = QDateEdit(QDate(QDate.currentDate().year()+1,6,30))
        self.dt_end.setCalendarPopup(True); self.dt_end.setFixedHeight(38)
        self.dt_end.setStyleSheet(self._is()); c4.addWidget(self.dt_end)
        r2.addLayout(c3,1); r2.addLayout(c4,1); fl.addLayout(r2)
        af = QFrame(); af.setStyleSheet("background:#EFF6FF; border-radius:6px; padding:8px;")
        al = QHBoxLayout(af)
        al.addWidget(QLabel("Export auto: chaque vendredi 18h -> Documents/Rapports_Academiques/"))
        bt = QPushButton("Tester")
        bt.setStyleSheet("background:#1F4E79; color:white; border-radius:5px; padding:3px 10px;")
        bt.clicked.connect(lambda: self._auto_export(False))
        al.addWidget(bt); fl.addWidget(af)
        self.prog_fr = QFrame(); self.prog_fr.setVisible(False)
        pl = QVBoxLayout(self.prog_fr)
        self.prog_lbl = QLabel("..."); self.prog_lbl.setStyleSheet("color:#6B7280; font-size:11px;")
        pl.addWidget(self.prog_lbl)
        self.prog_bar = QProgressBar()
        self.prog_bar.setStyleSheet("QProgressBar{border:1px solid #E5E7EB; border-radius:5px; height:18px;} QProgressBar::chunk{background:#10B981; border-radius:4px;}")
        pl.addWidget(self.prog_bar); fl.addWidget(self.prog_fr)
        hl = QHBoxLayout(); hl.addStretch()
        self.btn_gen = QPushButton("Generer le rapport")
        self.btn_gen.setFixedHeight(42)
        self.btn_gen.setStyleSheet("QPushButton{background:#1F4E79; color:white; border:none; border-radius:7px; padding:0 24px; font-size:13px; font-weight:bold;} QPushButton:hover{background:#2E75B6;} QPushButton:disabled{background:#9CA3AF;}")
        self.btn_gen.clicked.connect(self._generate)
        hl.addWidget(self.btn_gen); fl.addLayout(hl)
        lay.addWidget(fr); lay.addStretch()
        return w

    def _generate(self):
        fmt   = self.fmt_cb.currentText()
        rtype = self.rtype_cb.currentText()
        sd = self.dt_start.date(); ed = self.dt_end.date()
        ps = f"{sd.day():02d}/{sd.month():02d}/{sd.year()}"
        pe = f"{ed.day():02d}/{ed.month():02d}/{ed.year()}"
        ext = ".pdf" if "PDF" in fmt else ".xlsx"
        fn, _ = QFileDialog.getSaveFileName(self,"Enregistrer",
                    f"rapport_{datetime.now().strftime('%Y%m%d')}{ext}",
                    "PDF (*.pdf);;Excel (*.xlsx);;Tous (*.*)")
        if not fn: return
        self.btn_gen.setEnabled(False); self.btn_gen.setText("Generation...")
        self.prog_fr.setVisible(True); self.prog_bar.setValue(0)
        self.thread = ReportThread(fmt, fn, self.session, ps, pe, rtype)
        self.thread.progress.connect(lambda v,m: (self.prog_bar.setValue(v), self.prog_lbl.setText(m)))
        self.thread.done.connect(self._on_done)
        self.thread.error.connect(self._on_error)
        self.thread.start()

    def _on_done(self, r):
        self.btn_gen.setEnabled(True); self.btn_gen.setText("Generer le rapport")
        self.prog_fr.setVisible(False)
        QMessageBox.information(self,"Succes",f"Rapport genere !\n{r['filename']}")

    def _on_error(self, e):
        self.btn_gen.setEnabled(True); self.btn_gen.setText("Generer le rapport")
        self.prog_fr.setVisible(False)
        QMessageBox.critical(self,"Erreur",e[:400])

    def _is(self):
        return "QComboBox,QDateEdit{background:white; border:1px solid #D1D5DB; border-radius:5px; padding:0 8px; font-size:12px;}"

    def load_stats(self): pass
    def showEvent(self, event): super().showEvent(event)