"""
<<<<<<< HEAD
Onglet de gestion des enseignants — VERSION SQLite
Toutes les opérations utilisent SQLAlchemy / db_manager.
=======
Onglet de gestion des enseignants - VERSION CORRIGÉE
Toutes les fonctionnalités sont maintenant connectées
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget,
    QTableWidgetItem, QHeaderView, QLabel, QLineEdit, QComboBox,
    QMessageBox, QDialog, QFormLayout, QSpinBox, QFileDialog
)
from PyQt5.QtCore import Qt
<<<<<<< HEAD

from datetime import datetime
import json

from src.database.db_manager import db_manager
from src.database.repositories import TeacherRepository, UFRRepository, CohortRepository
from src.database.models import TeacherModel


class TeacherDialog(QDialog):
    """Dialogue pour ajouter/modifier un enseignant"""

    def __init__(self, parent=None, teacher_data=None, ufrs=None):
        super().__init__(parent)
        self.teacher_data = teacher_data or {}
        self.ufrs = ufrs or []
=======
from datetime import datetime
import json


class TeacherDialog(QDialog):
    """Dialogue pour ajouter/modifier un enseignant"""
    
    def __init__(self, parent=None, teacher_data=None):
        super().__init__(parent)
        self.teacher_data = teacher_data or {}
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.setWindowTitle("Enseignant" if not teacher_data else "Modifier Enseignant")
        self.setModal(True)
        self.setMinimumWidth(500)
        self.init_ui()
<<<<<<< HEAD

    def init_ui(self):
        """Initialiser l'interface du dialogue"""
        layout = QFormLayout(self)

=======
    
    def init_ui(self):
        """Initialiser l'interface du dialogue"""
        layout = QFormLayout(self)
        
        # Champs de saisie
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.nom_edit = QLineEdit(self.teacher_data.get('nom', ''))
        self.prenom_edit = QLineEdit(self.teacher_data.get('prenom', ''))
        self.email_edit = QLineEdit(self.teacher_data.get('email', ''))
        self.telephone_edit = QLineEdit(self.teacher_data.get('telephone', ''))
<<<<<<< HEAD
        self.specialite_edit = QLineEdit(self.teacher_data.get('specialite', ''))

=======
        
        self.specialite_edit = QLineEdit(self.teacher_data.get('specialite', ''))
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.statut_combo = QComboBox()
        self.statut_combo.addItems(["Permanent", "Vacataire", "Contractuel"])
        if 'statut' in self.teacher_data:
            idx = self.statut_combo.findText(self.teacher_data['statut'])
            if idx >= 0:
                self.statut_combo.setCurrentIndex(idx)
<<<<<<< HEAD

=======
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.heures_spin = QSpinBox()
        self.heures_spin.setRange(0, 60)
        self.heures_spin.setValue(self.teacher_data.get('heures_semaine', 40))
        self.heures_spin.setSuffix(" h/semaine")
<<<<<<< HEAD

        self.ufr_combo = QComboBox()
        self.ufr_combo.addItem("Aucune UFR", None)
        if self.ufrs:
            for ufr in self.ufrs:
                label = f"{ufr.get('nom', '')} ({ufr.get('code', '')})"
                self.ufr_combo.addItem(label, ufr.get('id'))
        else:
            self.ufr_combo.setEnabled(False)

        if 'ufr_id' in self.teacher_data and self.teacher_data['ufr_id']:
            for i in range(self.ufr_combo.count()):
                if self.ufr_combo.itemData(i) == self.teacher_data['ufr_id']:
                    self.ufr_combo.setCurrentIndex(i)
                    break

=======
        
        # Ajouter les champs au formulaire
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        layout.addRow("Nom:", self.nom_edit)
        layout.addRow("Prénom:", self.prenom_edit)
        layout.addRow("Email:", self.email_edit)
        layout.addRow("Téléphone:", self.telephone_edit)
        layout.addRow("Spécialité:", self.specialite_edit)
        layout.addRow("Statut:", self.statut_combo)
        layout.addRow("Heures/semaine:", self.heures_spin)
<<<<<<< HEAD
        layout.addRow("UFR de rattachement:", self.ufr_combo)

        btn_layout = QHBoxLayout()
        btn_save = QPushButton("💾 Enregistrer")
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50; color: white; border: none;
                border-radius: 4px; padding: 10px 20px; font-weight: bold;
            }
            QPushButton:hover { background-color: #45a049; }
        """)
        btn_save.clicked.connect(self.accept)

        btn_cancel = QPushButton("❌ Annuler")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #f44336; color: white; border: none;
                border-radius: 4px; padding: 10px 20px; font-weight: bold;
            }
            QPushButton:hover { background-color: #da190b; }
        """)
        btn_cancel.clicked.connect(self.reject)

        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)
        layout.addRow(btn_layout)

    def get_data(self):
=======
        
        # Boutons
        btn_layout = QHBoxLayout()
        
        btn_save = QPushButton("💾 Enregistrer")
        btn_save.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
        """)
        btn_save.clicked.connect(self.accept)
        
        btn_cancel = QPushButton("❌ Annuler")
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 10px 20px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
        """)
        btn_cancel.clicked.connect(self.reject)
        
        btn_layout.addWidget(btn_save)
        btn_layout.addWidget(btn_cancel)
        
        layout.addRow(btn_layout)
    
    def get_data(self):
        """Récupérer les données du formulaire"""
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        return {
            'nom': self.nom_edit.text().strip(),
            'prenom': self.prenom_edit.text().strip(),
            'email': self.email_edit.text().strip(),
            'telephone': self.telephone_edit.text().strip(),
            'specialite': self.specialite_edit.text().strip(),
            'statut': self.statut_combo.currentText(),
<<<<<<< HEAD
            'heures_semaine': self.heures_spin.value(),
            'ufr_id': self.ufr_combo.currentData()
        }

    def validate(self):
        data = self.get_data()
        if not data['nom']:
            QMessageBox.warning(self, "Validation", "Le nom est obligatoire !")
            return False
        if not data['prenom']:
            QMessageBox.warning(self, "Validation", "Le prénom est obligatoire !")
            return False
        if not data['email']:
            QMessageBox.warning(self, "Validation", "L'email est obligatoire !")
            return False
        if '@' not in data['email']:
            QMessageBox.warning(self, "Validation", "Email invalide !")
            return False
        return True

    def accept(self):
=======
            'heures_semaine': self.heures_spin.value()
        }
    
    def validate(self):
        """Valider les données"""
        data = self.get_data()
        
        if not data['nom']:
            QMessageBox.warning(self, "Validation", "Le nom est obligatoire !")
            return False
        
        if not data['prenom']:
            QMessageBox.warning(self, "Validation", "Le prénom est obligatoire !")
            return False
        
        if not data['email']:
            QMessageBox.warning(self, "Validation", "L'email est obligatoire !")
            return False
        
        if '@' not in data['email']:
            QMessageBox.warning(self, "Validation", "Email invalide !")
            return False
        
        return True
    
    def accept(self):
        """Accepter si validation OK"""
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        if self.validate():
            super().accept()


class TeachersTab(QWidget):
<<<<<<< HEAD
    """Onglet pour gérer les enseignants — VERSION SQLite"""

    def __init__(self, current_user=None, parent=None, activities_tab=None):
        super().__init__(parent)
        self.current_user = current_user
        self.activities_tab = activities_tab
        # ✅ SQLite via db_manager
        self.session = db_manager.get_session()
        self.teacher_repo = TeacherRepository(self.session)
        self.ufr_repo = UFRRepository(self.session)
        self.cohort_repo = CohortRepository(self.session)
        self.teachers = []
        self.filtered_teachers = []
        self.ufrs = []
        self.cohortes = []
        self.init_ui()
        self.load_ufrs()
        self.load_cohortes()
        self.load_teachers()

    # ==========================================
    # CHARGEMENT DEPUIS SQLITE
    # ==========================================

    def load_teachers(self):
        """Charge les enseignants depuis SQLite."""
        try:
            rows = self.teacher_repo.get_all()
            self.teachers = []
            for t in rows:
                parts = (t.full_name or '').split(' ', 1)
                nom = parts[0]
                prenom = parts[1] if len(parts) > 1 else ''
                ufr_nom = None
                if t.ufr_id:
                    ufr = self.ufr_repo.get_by_id(t.ufr_id)
                    if ufr:
                        ufr_nom = ufr.name
                self.teachers.append({
                    'id': t.id,
                    'nom': nom,
                    'prenom': prenom,
                    'email': t.email or '',
                    'telephone': t.phone or '',
                    'specialite': t.speciality or '',
                    'statut': t.status.value if (t.status and hasattr(t.status, 'value')) else (str(t.status) if t.status else ''),
                    'heures_semaine': t.max_hours_per_week or 0,
                    'ufr_id': t.ufr_id,
                    'ufr_nom': ufr_nom or 'N/A'
                })
            self.apply_filters()
        except Exception as e:
            print(f"Erreur chargement enseignants: {e}")
            self.teachers = []
            self.apply_filters()

    def load_ufrs(self):
        """Charge les UFR depuis SQLite."""
        try:
            rows = self.ufr_repo.get_all()
            self.ufrs = [
                {'id': u.id, 'nom': u.name, 'code': u.code}
                for u in rows
            ]
        except Exception as e:
            print(f"Erreur chargement UFR: {e}")
            self.ufrs = []

    def load_cohortes(self):
        """Charge les cohortes depuis SQLite."""
        try:
            rows = self.cohort_repo.get_all()
            self.cohortes = [
                {
                    'id': c.id,
                    'nom': c.name,
                    'annee_academique': c.academic_year or '',
                    'semestre': c.semester or '',
                    'effectif': c.student_count or 0
                }
                for c in rows
            ]
        except Exception as e:
            print(f"Erreur chargement cohortes: {e}")
            self.cohortes = []

    # ==========================================
    # INTERFACE
    # ==========================================

=======
    """Onglet pour gérer les enseignants - VERSION CORRIGÉE"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.teachers = []  # Liste des enseignants
        self.filtered_teachers = []  # Liste filtrée
        self.init_ui()
        self.load_sample_data()  # Charger des données exemple
    
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
    def init_ui(self):
        """Initialise l'interface."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(20)
<<<<<<< HEAD

=======
        
        # En-tête
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        header_layout = QVBoxLayout()
        title = QLabel("Enseignants")
        title.setStyleSheet("font-size: 32px; font-weight: bold; color: #1a1a1a;")
        header_layout.addWidget(title)
<<<<<<< HEAD
        subtitle = QLabel("Gérez les enseignants et leurs contraintes")
        subtitle.setStyleSheet("font-size: 16px; color: #666;")
        header_layout.addWidget(subtitle)
        layout.addLayout(header_layout)

        search_layout = QHBoxLayout()
=======
        
        subtitle = QLabel("Gérez les enseignants et leurs contraintes")
        subtitle.setStyleSheet("font-size: 16px; color: #666;")
        header_layout.addWidget(subtitle)
        
        layout.addLayout(header_layout)
        
        # Barre de recherche et filtres
        search_layout = QHBoxLayout()
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍 Rechercher un enseignant...")
        self.search_box.setStyleSheet("""
            QLineEdit {
<<<<<<< HEAD
                padding: 12px; border: 1px solid #E0E0E0;
                border-radius: 8px; font-size: 14px; background: white;
            }
        """)
        self.search_box.setFixedHeight(45)
        self.search_box.textChanged.connect(self.apply_filters)

        self.filter_box = QComboBox()
        self.filter_box.addItems(["Tous", "Permanents", "Vacataires", "Contractuels"])
        self.filter_box.setStyleSheet("""
            QComboBox {
                padding: 12px; border: 1px solid #E0E0E0;
                border-radius: 8px; background: white; min-width: 150px;
            }
        """)
        self.filter_box.setFixedHeight(45)
        self.filter_box.currentTextChanged.connect(self.apply_filters)

        search_layout.addWidget(self.search_box, 3)
        search_layout.addWidget(self.filter_box, 1)
        layout.addLayout(search_layout)

        btn_layout = QHBoxLayout()
        self.btn_add = QPushButton("➕ Nouvel Enseignant")
        self.btn_add.setStyleSheet(self.get_button_style("#4CAF50"))
        self.btn_add.setFixedHeight(40)
        self.btn_add.clicked.connect(self.add_teacher)

        self.btn_edit = QPushButton("✏️ Modifier")
        self.btn_edit.setStyleSheet(self.get_button_style("#2196F3"))
        self.btn_edit.setFixedHeight(40)
        self.btn_edit.clicked.connect(self.edit_teacher)

        self.btn_delete = QPushButton("🗑️ Supprimer")
        self.btn_delete.setStyleSheet(self.get_button_style("#F44336"))
        self.btn_delete.setFixedHeight(40)
        self.btn_delete.clicked.connect(self.delete_teacher)

        self.btn_export = QPushButton("📥 Exporter")
        self.btn_export.setStyleSheet(self.get_button_style("#FF9800"))
        self.btn_export.setFixedHeight(40)
        self.btn_export.clicked.connect(self.export_teachers)

=======
                padding: 12px;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                font-size: 14px;
                background: white;
            }
        """)
        self.search_box.setFixedHeight(45)
        self.search_box.textChanged.connect(self.apply_filters)  # ✅ CONNECTÉ
        
        self.filter_box = QComboBox()
        self.filter_box.addItems(["Tous", "Permanents", "Vacataires", "Contractuels", "Disponibles"])
        self.filter_box.setStyleSheet("""
            QComboBox {
                padding: 12px;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background: white;
                min-width: 150px;
            }
        """)
        self.filter_box.setFixedHeight(45)
        self.filter_box.currentTextChanged.connect(self.apply_filters)  # ✅ CONNECTÉ
        
        search_layout.addWidget(self.search_box, 3)
        search_layout.addWidget(self.filter_box, 1)
        
        layout.addLayout(search_layout)
        
        # Boutons d'action
        btn_layout = QHBoxLayout()
        
        self.btn_add = QPushButton("➕ Nouvel Enseignant")
        self.btn_add.setStyleSheet(self.get_button_style("#4CAF50"))
        self.btn_add.setFixedHeight(40)
        self.btn_add.clicked.connect(self.add_teacher)  # ✅ CONNECTÉ
        
        self.btn_edit = QPushButton("✏️ Modifier")
        self.btn_edit.setStyleSheet(self.get_button_style("#2196F3"))
        self.btn_edit.setFixedHeight(40)
        self.btn_edit.clicked.connect(self.edit_teacher)  # ✅ CONNECTÉ
        
        self.btn_delete = QPushButton("🗑️ Supprimer")
        self.btn_delete.setStyleSheet(self.get_button_style("#F44336"))
        self.btn_delete.setFixedHeight(40)
        self.btn_delete.clicked.connect(self.delete_teacher)  # ✅ CONNECTÉ
        
        self.btn_export = QPushButton("📥 Exporter")
        self.btn_export.setStyleSheet(self.get_button_style("#FF9800"))
        self.btn_export.setFixedHeight(40)
        self.btn_export.clicked.connect(self.export_teachers)  # ✅ CONNECTÉ
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        btn_layout.addWidget(self.btn_add)
        btn_layout.addWidget(self.btn_edit)
        btn_layout.addWidget(self.btn_delete)
        btn_layout.addWidget(self.btn_export)
        btn_layout.addStretch()
<<<<<<< HEAD
        layout.addLayout(btn_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels([
            "ID", "Nom", "Prénom", "Email", "Téléphone",
            "Spécialité", "Statut", "UFR", "Heures/semaine"
=======
        
        layout.addLayout(btn_layout)
        
        # Table des enseignants
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ID", "Nom", "Prénom", "Email", "Téléphone", 
            "Spécialité", "Statut", "Heures/semaine"
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setStyleSheet(self.get_table_style())
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
<<<<<<< HEAD
        layout.addWidget(self.table)

        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(20)
=======
        
        layout.addWidget(self.table)
        
        # Statistiques en bas
        self.stats_layout = QHBoxLayout()
        self.stats_layout.setSpacing(20)
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.stat1 = self.create_stat_box("Total enseignants", "0", "#2196F3")
        self.stat2 = self.create_stat_box("Permanents", "0", "#4CAF50")
        self.stat3 = self.create_stat_box("Vacataires", "0", "#FF9800")
        self.stat4 = self.create_stat_box("Heures totales", "0h", "#9C27B0")
<<<<<<< HEAD
        for s in [self.stat1, self.stat2, self.stat3, self.stat4]:
            self.stats_layout.addWidget(s)
        layout.addLayout(self.stats_layout)

    # ==========================================
    # FILTRES ET AFFICHAGE
    # ==========================================

    def apply_filters(self):
        search_text = self.search_box.text().lower()
        filter_status = self.filter_box.currentText()
        self.filtered_teachers = []
        for teacher in self.teachers:
=======
        
        self.stats_layout.addWidget(self.stat1)
        self.stats_layout.addWidget(self.stat2)
        self.stats_layout.addWidget(self.stat3)
        self.stats_layout.addWidget(self.stat4)
        
        layout.addLayout(self.stats_layout)
    
    def load_sample_data(self):
        """Charger des données exemple"""
        self.teachers = [
            {
                'id': 1,
                'nom': 'KABORE',
                'prenom': 'Marie',
                'email': 'marie.kabore@uts.bf',
                'telephone': '+226 70 12 34 56',
                'specialite': 'Algorithmique',
                'statut': 'Permanent',
                'heures_semaine': 40
            },
            {
                'id': 2,
                'nom': 'TRAORE',
                'prenom': 'Moussa',
                'email': 'moussa.traore@uts.bf',
                'telephone': '+226 70 23 45 67',
                'specialite': 'Bases de données',
                'statut': 'Permanent',
                'heures_semaine': 40
            },
            {
                'id': 3,
                'nom': 'SAWADOGO',
                'prenom': 'Fatimata',
                'email': 'fatimata.sawadogo@uts.bf',
                'telephone': '+226 70 34 56 78',
                'specialite': 'Réseaux',
                'statut': 'Permanent',
                'heures_semaine': 40
            },
            {
                'id': 4,
                'nom': 'OUATTARA',
                'prenom': 'Ibrahim',
                'email': 'ibrahim.ouattara@uts.bf',
                'telephone': '+226 70 45 67 89',
                'specialite': 'Développement Web',
                'statut': 'Vacataire',
                'heures_semaine': 20
            },
            {
                'id': 5,
                'nom': 'ZONGO',
                'prenom': 'Aminata',
                'email': 'aminata.zongo@uts.bf',
                'telephone': '+226 70 56 78 90',
                'specialite': 'Intelligence Artificielle',
                'statut': 'Permanent',
                'heures_semaine': 40
            },
        ]
        self.apply_filters()
    
    def apply_filters(self):
        """Appliquer les filtres de recherche"""
        search_text = self.search_box.text().lower()
        filter_status = self.filter_box.currentText()
        
        # Filtrer les enseignants
        self.filtered_teachers = []
        
        for teacher in self.teachers:
            # Filtre de recherche
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            if search_text:
                searchable = f"{teacher['nom']} {teacher['prenom']} {teacher['email']} {teacher['specialite']}".lower()
                if search_text not in searchable:
                    continue
<<<<<<< HEAD
=======
            
            # Filtre de statut
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            if filter_status != "Tous":
                if filter_status == "Permanents" and teacher['statut'] != "Permanent":
                    continue
                elif filter_status == "Vacataires" and teacher['statut'] != "Vacataire":
                    continue
                elif filter_status == "Contractuels" and teacher['statut'] != "Contractuel":
                    continue
<<<<<<< HEAD
            self.filtered_teachers.append(teacher)
        self.refresh_table()
        self.update_statistics()

    def refresh_table(self):
        self.table.setRowCount(len(self.filtered_teachers))
=======
            
            self.filtered_teachers.append(teacher)
        
        self.refresh_table()
        self.update_statistics()
    
    def refresh_table(self):
        """Rafraîchir l'affichage de la table"""
        self.table.setRowCount(len(self.filtered_teachers))
        
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        for i, teacher in enumerate(self.filtered_teachers):
            self.table.setItem(i, 0, QTableWidgetItem(str(teacher['id'])))
            self.table.setItem(i, 1, QTableWidgetItem(teacher['nom']))
            self.table.setItem(i, 2, QTableWidgetItem(teacher['prenom']))
            self.table.setItem(i, 3, QTableWidgetItem(teacher['email']))
            self.table.setItem(i, 4, QTableWidgetItem(teacher['telephone']))
            self.table.setItem(i, 5, QTableWidgetItem(teacher['specialite']))
<<<<<<< HEAD
=======
            
            # Colonne Statut avec couleur
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
            status_item = QTableWidgetItem(teacher['statut'])
            if teacher['statut'] == "Permanent":
                status_item.setBackground(Qt.green)
            elif teacher['statut'] == "Vacataire":
                status_item.setBackground(Qt.yellow)
            else:
                status_item.setBackground(Qt.cyan)
            self.table.setItem(i, 6, status_item)
<<<<<<< HEAD
            self.table.setItem(i, 7, QTableWidgetItem(teacher.get('ufr_nom', 'N/A')))
            self.table.setItem(i, 8, QTableWidgetItem(f"{teacher['heures_semaine']}h"))

    def update_statistics(self):
=======
            
            self.table.setItem(i, 7, QTableWidgetItem(f"{teacher['heures_semaine']}h"))
    
    def update_statistics(self):
        """Mettre à jour les statistiques"""
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        total = len(self.filtered_teachers)
        permanents = sum(1 for t in self.filtered_teachers if t['statut'] == 'Permanent')
        vacataires = sum(1 for t in self.filtered_teachers if t['statut'] == 'Vacataire')
        heures_totales = sum(t['heures_semaine'] for t in self.filtered_teachers)
<<<<<<< HEAD
=======
        
        # Mettre à jour les labels
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        self.update_stat_box(self.stat1, str(total))
        self.update_stat_box(self.stat2, str(permanents))
        self.update_stat_box(self.stat3, str(vacataires))
        self.update_stat_box(self.stat4, f"{heures_totales}h")
<<<<<<< HEAD

    def update_stat_box(self, box, value):
        labels = box.findChildren(QLabel)
        if labels:
            labels[0].setText(value)

    # ==========================================
    # ACTIONS CRUD — SQLite
    # ==========================================

    def add_teacher(self):
        """Ajouter un nouvel enseignant dans SQLite."""
        dialog = TeacherDialog(self, ufrs=self.ufrs)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            try:
                nom_complet = f"{data['nom']} {data['prenom']}".strip()
                teacher = TeacherModel(
                    full_name=nom_complet,
                    email=data.get('email', ''),
                    phone=data.get('telephone', ''),
                    speciality=data.get('specialite', ''),
                    max_hours_per_week=data.get('heures_semaine', 0),
                    ufr_id=data.get('ufr_id')
                )
                self.session.add(teacher)
                self.session.commit()
                self.load_teachers()
                if self.activities_tab is not None:
                    self.activities_tab.load_related_data()
                QMessageBox.information(
                    self, "Succès",
                    f"L'enseignant {data['prenom']} {data['nom']} a été ajouté avec succès !"
                )
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "Erreur", f"Impossible d'ajouter l'enseignant :\n{e}")

    def edit_teacher(self):
        """Modifier l'enseignant sélectionné dans SQLite."""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Aucune sélection", "Veuillez sélectionner un enseignant à modifier.")
            return
        row = selected_rows[0].row()
        teacher = self.filtered_teachers[row]
        dialog = TeacherDialog(self, teacher, self.ufrs)
        if dialog.exec_() == QDialog.Accepted:
            updated_data = dialog.get_data()
            try:
                t_model = self.teacher_repo.get_by_id(teacher['id'])
                if t_model:
                    t_model.full_name = f"{updated_data['nom']} {updated_data['prenom']}".strip()
                    t_model.email = updated_data.get('email', '')
                    t_model.phone = updated_data.get('telephone', '')
                    t_model.speciality = updated_data.get('specialite', '')
                    t_model.max_hours_per_week = updated_data.get('heures_semaine', 0)
                    t_model.ufr_id = updated_data.get('ufr_id')
                    self.session.commit()
                self.load_teachers()
                if self.activities_tab is not None:
                    self.activities_tab.load_related_data()
                QMessageBox.information(
                    self, "Succès",
                    f"L'enseignant {updated_data['prenom']} {updated_data['nom']} a été modifié avec succès !"
                )
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "Erreur", f"Impossible de modifier l'enseignant :\n{e}")

    def delete_teacher(self):
        """Supprimer l'enseignant sélectionné de SQLite."""
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Aucune sélection", "Veuillez sélectionner un enseignant à supprimer.")
            return
        row = selected_rows[0].row()
        teacher = self.filtered_teachers[row]
        reply = QMessageBox.question(
            self, "Confirmation",
            f"Êtes-vous sûr de vouloir supprimer l'enseignant {teacher['prenom']} {teacher['nom']} ?\n\n"
            "Cette action est irréversible.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            try:
                import sqlalchemy
                teacher_id = teacher['id']
                # 1. Supprimer les créneaux liés
                self.session.execute(
                    sqlalchemy.text("DELETE FROM schedule_slots WHERE teacher_id = :tid"),
                    {"tid": teacher_id}
                )
                # 2. Mettre teacher_id=NULL dans les activités
                self.session.execute(
                    sqlalchemy.text("UPDATE academic_activities SET teacher_id = NULL WHERE teacher_id = :tid"),
                    {"tid": teacher_id}
                )
                # 3. Supprimer l'enseignant
                self.session.execute(
                    sqlalchemy.text("DELETE FROM teachers WHERE id = :tid"),
                    {"tid": teacher_id}
                )
                self.session.commit()
                self.load_teachers()
                if self.activities_tab is not None:
                    self.activities_tab.load_related_data()
                QMessageBox.information(
                    self, "Succès",
                    f"L'enseignant {teacher['prenom']} {teacher['nom']} a été supprimé avec succès !"
                )
            except Exception as e:
                self.session.rollback()
                QMessageBox.critical(self, "Erreur", f"Impossible de supprimer l'enseignant :\n{e}")

    def export_teachers(self):
        """Exporter la liste des enseignants en JSON."""
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self, "Exporter les enseignants",
                f"enseignants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                "JSON Files (*.json)"
            )
            if filename:
                with open(filename, 'w', encoding='utf-8') as f:
                    json.dump(self.filtered_teachers, f, indent=2, ensure_ascii=False)
                QMessageBox.information(
                    self, "Succès",
                    f"Export réussi !\nFichier : {filename}\nEnseignants exportés : {len(self.filtered_teachers)}"
                )
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Erreur lors de l'export :\n{e}")

    # ==========================================
    # STYLES
    # ==========================================

    def create_stat_box(self, label, value, color):
        box = QWidget()
        box.setStyleSheet(f"QWidget {{ background-color: {color}; border-radius: 8px; padding: 15px; }}")
        layout = QVBoxLayout(box)
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: white;")
        value_label.setAlignment(Qt.AlignCenter)
        label_label = QLabel(label)
        label_label.setStyleSheet("font-size: 12px; color: white;")
        label_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(value_label)
        layout.addWidget(label_label)
        return box

    def get_button_style(self, color):
        return f"""
            QPushButton {{
                background-color: {color}; color: white; border: none;
                border-radius: 8px; padding: 10px 20px;
                font-size: 14px; font-weight: bold;
            }}
            QPushButton:hover {{ opacity: 0.9; }}
        """

    def get_table_style(self):
        return """
            QTableWidget {
                border: 1px solid #E0E0E0; border-radius: 8px;
                background-color: white; gridline-color: #F0F0F0;
            }
            QTableWidget::item { padding: 10px; }
            QTableWidget::item:selected { background-color: #E3F2FD; color: #1976D2; }
            QHeaderView::section {
                background-color: #F5F5F5; padding: 12px;
                border: none; font-weight: bold; color: #333;
            }
            QTableWidget::item:alternate { background-color: #FAFAFA; }
=======
    
    def update_stat_box(self, box, value):
        """Mettre à jour une boîte de statistique"""
        value_label = box.findChild(QLabel)
        if value_label:
            # Trouver le premier label (la valeur)
            labels = box.findChildren(QLabel)
            if labels:
                labels[0].setText(value)
    
    # ==========================================
    # MÉTHODES D'ACTION
    # ==========================================
    
    def add_teacher(self):
        """Ajouter un nouvel enseignant"""
        dialog = TeacherDialog(self)
        
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            
            # Générer un nouvel ID
            new_id = max([t['id'] for t in self.teachers], default=0) + 1
            data['id'] = new_id
            
            # Ajouter à la liste
            self.teachers.append(data)
            
            # Rafraîchir l'affichage
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"L'enseignant {data['prenom']} {data['nom']} a été ajouté avec succès !"
            )
    
    def edit_teacher(self):
        """Modifier l'enseignant sélectionné"""
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(
                self,
                "Aucune sélection",
                "Veuillez sélectionner un enseignant à modifier."
            )
            return
        
        # Récupérer l'enseignant sélectionné
        row = selected_rows[0].row()
        teacher = self.filtered_teachers[row]
        
        # Ouvrir le dialogue avec les données actuelles
        dialog = TeacherDialog(self, teacher)
        
        if dialog.exec_() == QDialog.Accepted:
            # Mettre à jour les données
            updated_data = dialog.get_data()
            teacher.update(updated_data)
            
            # Rafraîchir l'affichage
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"L'enseignant {teacher['prenom']} {teacher['nom']} a été modifié avec succès !"
            )
    
    def delete_teacher(self):
        """Supprimer l'enseignant sélectionné"""
        selected_rows = self.table.selectionModel().selectedRows()
        
        if not selected_rows:
            QMessageBox.warning(
                self,
                "Aucune sélection",
                "Veuillez sélectionner un enseignant à supprimer."
            )
            return
        
        # Récupérer l'enseignant sélectionné
        row = selected_rows[0].row()
        teacher = self.filtered_teachers[row]
        
        # Demander confirmation
        reply = QMessageBox.question(
            self,
            "Confirmation",
            f"Êtes-vous sûr de vouloir supprimer l'enseignant {teacher['prenom']} {teacher['nom']} ?\n\n"
            "Cette action est irréversible.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Supprimer de la liste principale
            self.teachers.remove(teacher)
            
            # Rafraîchir l'affichage
            self.apply_filters()
            
            QMessageBox.information(
                self,
                "Succès",
                f"L'enseignant {teacher['prenom']} {teacher['nom']} a été supprimé avec succès !"
            )
    
    def export_teachers(self):
        """Exporter la liste des enseignants"""
        # Demander le format d'export
        reply = QMessageBox.question(
            self,
            "Format d'export",
            "Choisissez le format d'export :\n\n"
            "Oui = Excel (.xlsx)\n"
            "Non = JSON (.json)",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply == QMessageBox.Yes:
            # Export Excel
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exporter en Excel",
                f"enseignants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if filename:
                self.export_to_excel(filename)
        else:
            # Export JSON
            filename, _ = QFileDialog.getSaveFileName(
                self,
                "Exporter en JSON",
                f"enseignants_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                "JSON Files (*.json)"
            )
            
            if filename:
                self.export_to_json(filename)
    
    def export_to_excel(self, filename):
        """Exporter en Excel"""
        try:
            # Note: Nécessite openpyxl
            # pip install openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Enseignants"
            
            # En-têtes
            headers = ["ID", "Nom", "Prénom", "Email", "Téléphone", "Spécialité", "Statut", "Heures/semaine"]
            ws.append(headers)
            
            # Style des en-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            
            # Données
            for teacher in self.filtered_teachers:
                ws.append([
                    teacher['id'],
                    teacher['nom'],
                    teacher['prenom'],
                    teacher['email'],
                    teacher['telephone'],
                    teacher['specialite'],
                    teacher['statut'],
                    teacher['heures_semaine']
                ])
            
            # Ajuster les colonnes
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(filename)
            
            QMessageBox.information(
                self,
                "Succès",
                f"Export Excel réussi !\n\nFichier : {filename}"
            )
            
        except ImportError:
            QMessageBox.warning(
                self,
                "Module manquant",
                "Le module 'openpyxl' n'est pas installé.\n\n"
                "Installez-le avec : pip install openpyxl"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export Excel :\n\n{str(e)}"
            )
    
    def export_to_json(self, filename):
        """Exporter en JSON"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.filtered_teachers, f, indent=2, ensure_ascii=False)
            
            QMessageBox.information(
                self,
                "Succès",
                f"Export JSON réussi !\n\nFichier : {filename}\nEnseignants exportés : {len(self.filtered_teachers)}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Erreur",
                f"Erreur lors de l'export JSON :\n\n{str(e)}"
            )
    
    # ==========================================
    # STYLES
    # ==========================================
    
    def create_stat_box(self, label, value, color):
        """Crée une petite boîte de statistique."""
        box = QWidget()
        box.setStyleSheet(f"""
            QWidget {{
                background-color: {color};
                border-radius: 8px;
                padding: 15px;
            }}
        """)
        
        layout = QVBoxLayout(box)
        
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold; color: white;")
        value_label.setAlignment(Qt.AlignCenter)
        
        label_label = QLabel(label)
        label_label.setStyleSheet("font-size: 12px; color: white;")
        label_label.setAlignment(Qt.AlignCenter)
        
        layout.addWidget(value_label)
        layout.addWidget(label_label)
        
        return box
    
    def get_button_style(self, color):
        """Style des boutons."""
        return f"""
            QPushButton {{
                background-color: {color};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                opacity: 0.9;
                background-color: {color};
            }}
            QPushButton:pressed {{
                background-color: {color};
            }}
        """
    
    def get_table_style(self):
        """Style de la table."""
        return """
            QTableWidget {
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                background-color: white;
                gridline-color: #F0F0F0;
            }
            QTableWidget::item {
                padding: 10px;
            }
            QTableWidget::item:selected {
                background-color: #E3F2FD;
                color: #1976D2;
            }
            QHeaderView::section {
                background-color: #F5F5F5;
                padding: 12px;
                border: none;
                font-weight: bold;
                color: #333;
            }
            QTableWidget::item:alternate {
                background-color: #FAFAFA;
            }
>>>>>>> a5a03a993e1b9b43f14c093746cbd6265ba0f65f
        """