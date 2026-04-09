"""
UC9 - Onglet Consultation des Emplois du Temps
100% SQLite : lit ScheduleSlotModel au lieu de schedules.json

CORRECTIONS :
- _reload_all et _populate_target_combo préservent la sélection actuelle
- Boutons CM / TD / TP / Examen sont désormais des filtres cliquables
  (ancienne version : simples labels décoratifs sans action)
"""
import csv
from datetime import date, datetime, timedelta
from datetime import time as dtime
from PyQt5.QtCore import QDate

from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QComboBox, QFileDialog, QMessageBox, QFrame,
    QScrollArea, QGridLayout, QSizePolicy,
    QDialog, QFormLayout, QTimeEdit, QDateEdit, QLineEdit, QMenu, QAction
)
from PyQt5.QtCore import Qt, QSize, QTime, pyqtSignal
from PyQt5.QtGui import QColor, QCursor

from src.database.db_manager import db_manager
from src.database.models import ScheduleSlotModel
from src.database.repositories.cohort_repository import CohortRepository
from src.database.repositories.teacher_repository import TeacherRepository


JOURS     = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi"]
HEURE_DEBUT = 8
HEURE_FIN   = 19
SLOT_H      = 65  # pixels par heure

TYPE_COLORS = {
    "CM":         ("#BBDEFB", "#1565C0"),
    "TD":         ("#C8E6C9", "#2E7D32"),
    "TP":         ("#FFE0B2", "#E65100"),
    "Examen":     ("#FCE4EC", "#880E4F"),
    "Soutenance": ("#E1BEE7", "#6A1B9A"),
}


def _alpha_to_border(alpha):
    if alpha is None:   return "#9E9E9E"
    if alpha >= 1.0:    return "#C62828"
    if alpha >= 0.5:    return "#F9A825"
    return "#2E7D32"


class CourseBlock(QFrame):
    edit_requested   = pyqtSignal(dict)
    delete_requested = pyqtSignal(dict)

    def __init__(self, slot_data, edit_mode=False, parent=None):
        super().__init__(parent)
        self.slot_data = slot_data
        self._edit_mode = edit_mode
        bg, fg       = TYPE_COLORS.get(slot_data.get('activity_type', ''), ("#E0E0E0", "#333"))
        border_color = _alpha_to_border(slot_data.get('alpha'))
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {bg};
                border: 1px solid #ccc;
                border-left: 4px solid {border_color};
                border-radius: 4px;
                margin: 1px;
            }}
        """)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(7, 5, 5, 5)
        layout.setSpacing(2)

        start    = slot_data.get('start', '08:00')
        end      = slot_data.get('end',   '10:00')
        hour_lbl = QLabel(f"⏰ {start} – {end}")
        hour_lbl.setStyleSheet(
            f"color: {fg}; font-weight: bold; font-size: 11px; "
            f"background: rgba(0,0,0,0.07); border-radius:3px; padding:1px 4px;"
        )
        layout.addWidget(hour_lbl)

        name_lbl = QLabel(slot_data.get('activity', '—'))
        name_lbl.setStyleSheet(f"color: {fg}; font-weight: bold; font-size: 11px;")
        name_lbl.setWordWrap(True)
        layout.addWidget(name_lbl)

        act_type = slot_data.get('activity_type', '')
        room     = slot_data.get('room', '')
        info_lbl = QLabel(f"📍 {act_type}  |  {room}")
        info_lbl.setStyleSheet(f"color: {fg}; font-size: 10px;")
        layout.addWidget(info_lbl)

        teach_lbl = QLabel(f"👤 {slot_data.get('teacher', '')}")
        teach_lbl.setStyleSheet("color: #444; font-size: 10px;")
        layout.addWidget(teach_lbl)

        self.setToolTip(
            f"Activité : {slot_data.get('activity', '')}\n"
            f"Type     : {act_type}\n"
            f"Salle    : {room}\n"
            f"Ens.     : {slot_data.get('teacher', '')}\n"
            f"Cohorte  : {slot_data.get('cohort', '')}\n"
            f"α        : {slot_data.get('alpha', 'N/A')}"
        )
        if self._edit_mode:
            self.setCursor(Qt.PointingHandCursor)
            self._apply_edit_border()
            self.setContextMenuPolicy(Qt.CustomContextMenu)
            self.customContextMenuRequested.connect(self._show_context_menu)

    def _apply_edit_border(self):
        self.setStyleSheet(self.styleSheet() +
            "QFrame { outline: 2px dashed #F59E0B; }"
        )

    def _show_context_menu(self, pos):
        menu = QMenu(self)
        act_edit   = menu.addAction("✏️  Modifier ce créneau")
        menu.addSeparator()
        act_delete = menu.addAction("🗑️  Supprimer ce créneau")
        menu.setStyleSheet("QMenu::item { padding: 6px 20px; } ")
        chosen = menu.exec_(QCursor.pos())
        if chosen == act_edit:
            self.edit_requested.emit(self.slot_data)
        elif chosen == act_delete:
            self.delete_requested.emit(self.slot_data)


# ══════════════════════════════════════════════════════════════════
# DIALOGUE D'ÉDITION D'UN CRÉNEAU
# ══════════════════════════════════════════════════════════════════

class SlotEditDialog(QDialog):
    def __init__(self, slot_data, teachers, parent=None):
        super().__init__(parent)
        self.slot_data  = slot_data
        self.teachers   = teachers
        self.setWindowTitle("Modifier le créneau")
        self.setMinimumWidth(420)
        self.setModal(True)
        self._build()

    def _build(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(14)

        title = QLabel(f"✏️  {self.slot_data.get('activity', '?')}")
        title.setStyleSheet("font-size:15px; font-weight:bold; color:#1a73e8;")
        layout.addWidget(title)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignRight)
        st = "border:1px solid #ddd; border-radius:5px; padding:0 8px; height:34px;"

        self._date = QDateEdit()
        self._date.setCalendarPopup(True)
        self._date.setStyleSheet(st)
        try:
            d = self.slot_data.get('date')
            if isinstance(d, str):
                d = datetime.fromisoformat(d).date()
            self._date.setDate(QDate(d.year, d.month, d.day))
        except Exception:
            self._date.setDate(QDate.currentDate())
        form.addRow("Date :", self._date)

        self._start = QTimeEdit()
        self._start.setDisplayFormat("HH:mm")
        self._start.setStyleSheet(st)
        try:
            h, m = map(int, self.slot_data.get('start','08:00').split(':'))
            self._start.setTime(QTime(h, m))
        except Exception:
            self._start.setTime(QTime(8, 0))
        form.addRow("Heure début :", self._start)

        self._end = QTimeEdit()
        self._end.setDisplayFormat("HH:mm")
        self._end.setStyleSheet(st)
        try:
            h, m = map(int, self.slot_data.get('end','10:00').split(':'))
            self._end.setTime(QTime(h, m))
        except Exception:
            self._end.setTime(QTime(10, 0))
        form.addRow("Heure fin :", self._end)

        self._room = QLineEdit(self.slot_data.get('room', ''))
        self._room.setStyleSheet(st)
        self._room.setPlaceholderText("ex: AMPHI 750, Salle B201")
        form.addRow("Salle :", self._room)

        self._teacher_combo = QComboBox()
        self._teacher_combo.setStyleSheet(st)
        current_tid = self.slot_data.get('teacher_id')
        for t in self.teachers:
            self._teacher_combo.addItem(t.full_name, t.id)
            if t.id == current_tid:
                self._teacher_combo.setCurrentIndex(self._teacher_combo.count() - 1)
        form.addRow("Enseignant :", self._teacher_combo)

        layout.addLayout(form)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("Annuler")
        btn_cancel.setFixedHeight(38)
        btn_cancel.setStyleSheet(
            "border:1px solid #ddd; border-radius:6px; padding:0 16px; color:#555;"
        )
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)

        btn_save = QPushButton("💾  Enregistrer")
        btn_save.setFixedHeight(38)
        btn_save.setStyleSheet(
            "background:#1a73e8; color:white; border-radius:6px; font-weight:bold; padding:0 18px;"
        )
        btn_save.clicked.connect(self._validate_and_accept)
        btn_row.addWidget(btn_save)
        layout.addLayout(btn_row)

    def _validate_and_accept(self):
        if self._start.time() >= self._end.time():
            QMessageBox.warning(self, "Erreur",
                "L'heure de fin doit être après l'heure de début.")
            return
        if not self._room.text().strip():
            QMessageBox.warning(self, "Erreur", "La salle ne peut pas être vide.")
            return
        self.accept()

    def get_values(self):
        qd   = self._date.date()
        qt_s = self._start.time()
        qt_e = self._end.time()
        return {
            'date':       date(qd.year(), qd.month(), qd.day()),
            'start':      f"{qt_s.hour():02d}:{qt_s.minute():02d}",
            'end':        f"{qt_e.hour():02d}:{qt_e.minute():02d}",
            'room':       self._room.text().strip(),
            'teacher_id': self._teacher_combo.currentData(),
            'teacher':    self._teacher_combo.currentText(),
        }


class TimetableGrid(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        main = QHBoxLayout(self)
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)

        time_col = QWidget()
        time_col.setFixedWidth(55)
        tcl = QVBoxLayout(time_col)
        tcl.setContentsMargins(0, 30, 0, 0)
        tcl.setSpacing(0)
        for h in range(HEURE_DEBUT, HEURE_FIN):
            lbl = QLabel(f"{h:02d}:00")
            lbl.setFixedHeight(SLOT_H)
            lbl.setStyleSheet("color: #555; font-size: 10px;")
            lbl.setAlignment(Qt.AlignTop | Qt.AlignRight)
            tcl.addWidget(lbl)
        tcl.addStretch()
        main.addWidget(time_col)

        self.grid_widget = QWidget()
        self.grid_layout = QGridLayout(self.grid_widget)
        self.grid_layout.setContentsMargins(0, 0, 0, 0)
        self.grid_layout.setSpacing(2)

        for col, jour in enumerate(JOURS):
            h = QLabel(jour)
            h.setFixedHeight(28)
            h.setAlignment(Qt.AlignCenter)
            h.setStyleSheet(
                "font-weight: bold; font-size: 12px; color: #1a73e8; "
                "background: #e8f0fe; border-radius: 4px;"
            )
            self.grid_layout.addWidget(h, 0, col)

        for row in range(1, HEURE_FIN - HEURE_DEBUT + 1):
            for col in range(len(JOURS)):
                cell = QFrame()
                cell.setFixedHeight(SLOT_H)
                cell.setStyleSheet("background: #fafafa; border: 1px solid #eee;")
                self.grid_layout.addWidget(cell, row, col)

        self.grid_widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main.addWidget(self.grid_widget)

    def load_slots(self, slots, edit_mode=False, on_edit=None, on_delete=None):
        for i in reversed(range(self.grid_layout.count())):
            item = self.grid_layout.itemAt(i)
            if item and isinstance(item.widget(), CourseBlock):
                item.widget().deleteLater()
                self.grid_layout.removeItem(item)

        for slot in slots:
            try:
                slot_date = slot.get('date')
                if isinstance(slot_date, str):
                    slot_date = datetime.fromisoformat(slot_date).date()
                weekday = slot_date.weekday()
                if weekday > 5:
                    continue
                start_str = slot.get('start', '08:00')
                end_str   = slot.get('end',   '10:00')
                start_h   = int(start_str.split(':')[0]) if ':' in start_str else int(start_str)
                end_h     = int(end_str.split(':')[0])   if ':' in end_str   else int(end_str)
                if start_h < HEURE_DEBUT or end_h > HEURE_FIN:
                    continue
                row_start = start_h - HEURE_DEBUT + 1
                row_span  = max(1, end_h - start_h)
                block = CourseBlock(slot, edit_mode=edit_mode)
                block.setFixedHeight(SLOT_H * row_span - 4)
                if edit_mode:
                    if on_edit:
                        block.edit_requested.connect(on_edit)
                    if on_delete:
                        block.delete_requested.connect(on_delete)
                self.grid_layout.addWidget(block, row_start, weekday, row_span, 1)
            except Exception:
                pass


class TimetableTab(QWidget):
    """UC9 — Consultation des emplois du temps — 100% SQLite."""

    def __init__(self, current_user=None, parent=None):
        super().__init__(parent)
        self.current_user = current_user
        self._all_slots   = []
        self._teachers    = []
        self._cohorts     = []
        self._edit_mode   = False
        self._type_filter = None   # ✅ Filtre par type (CM/TD/TP/Examen)
        self._type_btns   = {}     # ✅ Référence aux boutons de filtre
        today = date.today()
        self._week_start = today - timedelta(days=today.weekday())
        self._init_ui()
        self._load_db_data()
        self._update_week_label()

    def _load_db_data(self):
        try:
            session = db_manager.get_session()
            self._teachers = TeacherRepository(session).get_all()
            self._cohorts  = CohortRepository(session).get_all()
            self._load_slots_from_db(session)
            self._populate_target_combo()
        except Exception as e:
            print(f"[timetable_tab] Erreur DB: {e}")

    def _load_slots_from_db(self, session):
        self._all_slots = []
        try:
            slots_orm = session.query(ScheduleSlotModel).all()
            for s in slots_orm:
                if not s.date:
                    continue
                act_type = ""
                if s.activity and s.activity.type:
                    act_type = (s.activity.type.value
                                if hasattr(s.activity.type, 'value')
                                else str(s.activity.type))
                alpha = None
                if s.notes and 'alpha=' in s.notes:
                    try:
                        alpha = float(s.notes.split('alpha=')[1].split()[0])
                    except Exception:
                        alpha = s.delay_value
                start_str = (s.start_time.strftime('%H:%M')
                             if isinstance(s.start_time, dtime)
                             else str(s.start_time)[:5])
                end_str   = (s.end_time.strftime('%H:%M')
                             if isinstance(s.end_time, dtime)
                             else str(s.end_time)[:5])

                # ✅ Normaliser le type pour correspondre aux clés de TYPE_COLORS
                normalized_type = self._normalize_type(act_type)

                self._all_slots.append({
                    'date':          s.date,
                    'start':         start_str,
                    'end':           end_str,
                    'activity':      s.activity.name  if s.activity  else "?",
                    'activity_code': s.activity.code  if s.activity  else "",
                    'activity_type': normalized_type,
                    'teacher':       s.teacher.full_name if s.teacher else "Non assigné",
                    'cohort':        s.cohort.name    if s.cohort    else "?",
                    'room':          s.room or "—",
                    'alpha':         alpha,
                    'delay':         s.delay_value,
                    'cohort_id':     s.cohort_id,
                    'teacher_id':    s.teacher_id,
                    'activity_id':   s.activity_id,
                    'slot_id':       s.id,
                })
        except Exception as e:
            print(f"[timetable_tab] Erreur lecture schedule_slots: {e}")

    def _normalize_type(self, raw_type: str) -> str:
        """
        Normalise le type d'activité pour correspondre aux clés de TYPE_COLORS.
        Ex: "Travaux Dirigés" → "TD", "Cours Magistral" → "CM", etc.
        """
        if not raw_type:
            return ""
        t = raw_type.strip().upper()
        # Correspondances directes
        if t in ("CM", "TD", "TP", "EXAMEN", "SOUTENANCE"):
            return t.capitalize() if t == "EXAMEN" else t
        # Correspondances par mots-clés
        if "MAGISTRAL" in t or "COURS" in t:
            return "CM"
        if "DIRIGÉ" in t or "DIRIGE" in t or "TD" in t:
            return "TD"
        if "PRATIQUE" in t or "TP" in t:
            return "TP"
        if "EXAMEN" in t:
            return "Examen"
        if "SOUTENANCE" in t:
            return "Soutenance"
        # Retourner tel quel si non reconnu
        return raw_type

    def _on_type_filter(self, type_label: str, checked: bool):
        """
        Active/désactive le filtre par type d'activité.
        Un seul type peut être actif à la fois.
        Si on reclique sur le bouton déjà actif → réinitialise le filtre.
        """
        if checked:
            self._type_filter = type_label
            # Désactiver les autres boutons sans déclencher leur signal
            for lbl, btn in self._type_btns.items():
                if lbl != type_label:
                    btn.blockSignals(True)
                    btn.setChecked(False)
                    btn.blockSignals(False)
        else:
            self._type_filter = None
        self._refresh_grid()

    def _populate_target_combo(self):
        prev_data = self.target_combo.currentData()
        self.target_combo.blockSignals(True)
        self.target_combo.clear()
        view = self.view_combo.currentText()

        if view == "Par cohorte":
            if self._cohorts:
                for c in self._cohorts:
                    self.target_combo.addItem(
                        f"{c.name} ({c.academic_year})",
                        userData=('cohort', c.id, c.name)
                    )
            else:
                self.target_combo.addItem("⚠️ Aucune cohorte", userData=None)
        else:
            if self._teachers:
                for t in self._teachers:
                    self.target_combo.addItem(
                        t.full_name,
                        userData=('teacher', t.id, t.full_name)
                    )
            else:
                self.target_combo.addItem("⚠️ Aucun enseignant", userData=None)

        if prev_data:
            kind, entity_id, _ = prev_data
            for i in range(self.target_combo.count()):
                data = self.target_combo.itemData(i)
                if data and data[0] == kind and data[1] == entity_id:
                    self.target_combo.setCurrentIndex(i)
                    break

        self.target_combo.blockSignals(False)
        self._refresh_grid()

    def _reload_all(self):
        prev_data = self.target_combo.currentData()
        prev_view = self.view_combo.currentText()

        try:
            session = db_manager.get_session()
            self._teachers = TeacherRepository(session).get_all()
            self._cohorts  = CohortRepository(session).get_all()
            self._load_slots_from_db(session)
        except Exception as e:
            print(f"[timetable_tab] Erreur reload: {e}")

        self.view_combo.blockSignals(True)
        self.view_combo.setCurrentText(prev_view)
        self.view_combo.blockSignals(False)

        self.target_combo.blockSignals(True)
        self.target_combo.clear()

        if prev_view == "Par cohorte":
            for c in self._cohorts:
                self.target_combo.addItem(
                    f"{c.name} ({c.academic_year})",
                    userData=('cohort', c.id, c.name)
                )
        else:
            for t in self._teachers:
                self.target_combo.addItem(
                    t.full_name,
                    userData=('teacher', t.id, t.full_name)
                )

        if prev_data:
            kind, entity_id, _ = prev_data
            for i in range(self.target_combo.count()):
                data = self.target_combo.itemData(i)
                if data and data[0] == kind and data[1] == entity_id:
                    self.target_combo.setCurrentIndex(i)
                    break

        self.target_combo.blockSignals(False)
        self._refresh_grid()

    def _on_view_changed(self):
        self._populate_target_combo()

    def _update_week_label(self):
        end = self._week_start + timedelta(days=5)
        self.week_label.setText(
            f"{self._week_start.strftime('%d/%m')} — {end.strftime('%d/%m/%Y')}"
        )

    def _prev_week(self):
        self._week_start -= timedelta(weeks=1)
        self._update_week_label()
        self._refresh_grid()

    def _next_week(self):
        self._week_start += timedelta(weeks=1)
        self._update_week_label()
        self._refresh_grid()

    def _refresh_grid(self):
        target_data = self.target_combo.currentData()
        if not target_data or not self._all_slots:
            self.grid.load_slots([])
            self.empty_label.setVisible(True)
            self.scroll.setVisible(False)
            return

        self.empty_label.setVisible(False)
        self.scroll.setVisible(True)

        kind, entity_id, entity_name = target_data
        if kind == 'cohort':
            filtered = [s for s in self._all_slots
                        if s.get('cohort') == entity_name
                        or s.get('cohort_id') == entity_id]
        else:
            filtered = [s for s in self._all_slots
                        if s.get('teacher') == entity_name
                        or s.get('teacher_id') == entity_id]

        # ─── Filtrer par semaine affichée ────────────────────────
        week_end   = self._week_start + timedelta(days=6)
        week_slots = []
        for s in filtered:
            d = s.get('date')
            if isinstance(d, str):
                d = datetime.fromisoformat(d).date()
            if d is not None and self._week_start <= d <= week_end:
                week_slots.append(s)

        # ─── ✅ Filtrer par type (CM / TD / TP / Examen) ─────────
        if self._type_filter:
            week_slots = [
                s for s in week_slots
                if s.get('activity_type', '') == self._type_filter
            ]

        self.grid.load_slots(
            week_slots,
            edit_mode=self._edit_mode,
            on_edit=self._on_edit_slot,
            on_delete=self._on_delete_slot
        )

        # Mettre à jour le label d'info
        total_slots = len(filtered)
        type_info   = f" [{self._type_filter}]" if self._type_filter else ""
        if total_slots > 0 and len(week_slots) == 0:
            self.empty_label.setText(
                f"⚠️ Aucun cours{type_info} cette semaine.\n"
                f"({total_slots} créneau(x) au total — naviguez avec ◀ ▶)"
            )
            self.empty_label.setVisible(True)
            self.scroll.setVisible(True)
        elif total_slots == 0:
            self.empty_label.setText(
                "⚠️ Aucun emploi du temps disponible.\n"
                "Lancez d'abord l'ordonnancement (UC6)."
            )
            self.empty_label.setVisible(True)
            self.scroll.setVisible(False)

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        title = QLabel("📅 Emplois du Temps")
        title.setStyleSheet("font-size: 22px; font-weight: bold; color: #1a1a1a;")
        layout.addWidget(title)

        subtitle = QLabel("UC9 — Données lues depuis SQLite (schedule_slots)")
        subtitle.setStyleSheet("color: #555; font-size: 12px; font-style: italic;")
        layout.addWidget(subtitle)

        ctrl = QHBoxLayout()
        ctrl.setSpacing(10)
        ctrl.addWidget(QLabel("Vue :"))
        self.view_combo = QComboBox()
        self.view_combo.setFixedHeight(36)
        self.view_combo.addItems(["Par cohorte", "Par enseignant"])
        self.view_combo.currentIndexChanged.connect(self._on_view_changed)
        ctrl.addWidget(self.view_combo)

        ctrl.addWidget(QLabel("Afficher :"))
        self.target_combo = QComboBox()
        self.target_combo.setFixedHeight(36)
        self.target_combo.setMinimumWidth(250)
        self.target_combo.currentIndexChanged.connect(self._refresh_grid)
        ctrl.addWidget(self.target_combo)

        btn_reload = QPushButton("🔄 Recharger")
        btn_reload.setFixedHeight(36)
        btn_reload.setStyleSheet(
            "background: #e8f0fe; color: #1a73e8; border-radius: 6px; "
            "font-weight: bold; padding: 0 12px;"
        )
        btn_reload.clicked.connect(self._reload_all)
        ctrl.addWidget(btn_reload)

        self.btn_edit_mode = QPushButton("✏️  Mode édition")
        self.btn_edit_mode.setFixedHeight(36)
        self.btn_edit_mode.setCheckable(True)
        self.btn_edit_mode.setStyleSheet("""
            QPushButton {
                background: #F9FAFB; color: #374151;
                border: 1px solid #D1D5DB; border-radius: 6px;
                font-weight: bold; padding: 0 14px;
            }
            QPushButton:checked {
                background: #FEF3C7; color: #92400E;
                border: 1px solid #F59E0B;
            }
            QPushButton:hover { background: #F3F4F6; }
        """)
        self.btn_edit_mode.toggled.connect(self._toggle_edit_mode)
        ctrl.addWidget(self.btn_edit_mode)

        ctrl.addStretch()

        btn_prev = QPushButton("◀ Semaine préc.")
        btn_prev.setFixedHeight(36)
        btn_prev.setStyleSheet("background: #f5f5f5; border-radius:6px; padding:0 10px;")
        btn_prev.clicked.connect(self._prev_week)
        ctrl.addWidget(btn_prev)

        self.week_label = QLabel()
        self.week_label.setAlignment(Qt.AlignCenter)
        self.week_label.setMinimumWidth(160)
        self.week_label.setStyleSheet("font-weight:bold; font-size:12px;")
        ctrl.addWidget(self.week_label)

        btn_next = QPushButton("Semaine suiv. ▶")
        btn_next.setFixedHeight(36)
        btn_next.setStyleSheet("background: #f5f5f5; border-radius:6px; padding:0 10px;")
        btn_next.clicked.connect(self._next_week)
        ctrl.addWidget(btn_next)

        # ✅ Boutons de filtre CM / TD / TP / Examen — cliquables
        for label, (bg, fg) in list(TYPE_COLORS.items())[:4]:
            btn = QPushButton(label)
            btn.setFixedHeight(30)
            btn.setCheckable(True)
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {bg};
                    color: {fg};
                    border: 2px solid transparent;
                    border-radius: 4px;
                    font-size: 11px;
                    font-weight: bold;
                    padding: 2px 10px;
                }}
                QPushButton:checked {{
                    border: 2px solid {fg};
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    border: 1px solid {fg};
                }}
            """)
            btn.toggled.connect(
                lambda checked, t=label: self._on_type_filter(t, checked)
            )
            ctrl.addWidget(btn)
            self._type_btns[label] = btn

        layout.addLayout(ctrl)

        urg = QHBoxLayout()
        urg.setSpacing(5)
        urg.addWidget(QLabel("Bordure :"))
        for color, label in [
            ("#C62828", "🔴 Critique α≥1"),
            ("#F9A825", "🟡 Urgent 0.5≤α<1"),
            ("#2E7D32", "🟢 OK α<0.5"),
        ]:
            lbl = QLabel(label)
            lbl.setStyleSheet(
                f"border-left: 4px solid {color}; padding-left: 5px; "
                f"font-size: 11px; color: #555;"
            )
            urg.addWidget(lbl)
        urg.addStretch()
        layout.addLayout(urg)

        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.grid = TimetableGrid()
        self.scroll.setWidget(self.grid)
        self.scroll.setStyleSheet("border: 1px solid #ddd; border-radius: 8px;")
        layout.addWidget(self.scroll)

        self.empty_label = QLabel(
            "⚠️ Aucun emploi du temps disponible.\n"
            "Lancez d'abord l'ordonnancement (UC6)."
        )
        self.empty_label.setAlignment(Qt.AlignCenter)
        self.empty_label.setStyleSheet("color: #888; font-size: 14px;")
        self.empty_label.setVisible(False)
        layout.addWidget(self.empty_label)

        export_layout = QHBoxLayout()
        export_layout.setSpacing(8)
        for icon, text, handler, bg in [
            ("🖨️",  "Export PDF",   self._export_pdf,   "#c62828"),
            ("📊",  "Export Excel", self._export_excel, "#2e7d32"),
            ("📅",  "Export iCal",  self._export_ical,  "#1a73e8"),
            ("📄",  "Export CSV",   self._export_csv,   "#5e35b1"),
        ]:
            btn = QPushButton(f"{icon} {text}")
            btn.setFixedHeight(38)
            btn.setStyleSheet(
                f"QPushButton {{ background: {bg}; color: white; border-radius: 6px; "
                f"font-size: 13px; padding: 0 12px; }}"
                f"QPushButton:hover {{ opacity: 0.9; }}"
            )
            btn.clicked.connect(handler)
            export_layout.addWidget(btn)
        export_layout.addStretch()
        layout.addLayout(export_layout)

    def _get_current_slots(self):
        target_data = self.target_combo.currentData()
        if not target_data:
            return []
        kind, entity_id, entity_name = target_data
        if kind == 'cohort':
            return [s for s in self._all_slots
                    if s.get('cohort') == entity_name
                    or s.get('cohort_id') == entity_id]
        else:
            return [s for s in self._all_slots
                    if s.get('teacher') == entity_name
                    or s.get('teacher_id') == entity_id]

    def _export_pdf(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export PDF", "emploi_du_temps.pdf", "PDF (*.pdf)"
        )
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            doc = SimpleDocTemplate(path, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = []
            target_data = self.target_combo.currentData()
            title_txt = target_data[2] if target_data else "Emploi du temps"
            elements.append(Paragraph(f"Emploi du temps — {title_txt}", styles['Title']))
            elements.append(Spacer(1, 10))
            slots = self._get_current_slots()
            if not slots:
                elements.append(Paragraph("Aucun créneau disponible.", styles['Normal']))
            else:
                rows = [["Date", "Début", "Fin", "Activité", "Type", "Salle", "Enseignant"]]
                for s in sorted(slots, key=lambda x: (str(x.get('date', '')), x.get('start', ''))):
                    rows.append([str(s.get('date', '')), s.get('start', ''), s.get('end', ''),
                                 s.get('activity', ''), s.get('activity_type', ''),
                                 s.get('room', ''), s.get('teacher', '')])
                t = Table(rows)
                t.setStyle(TableStyle([
                    ('BACKGROUND',     (0,0), (-1,0), colors.HexColor('#1a73e8')),
                    ('TEXTCOLOR',      (0,0), (-1,0), colors.white),
                    ('FONTNAME',       (0,0), (-1,0), 'Helvetica-Bold'),
                    ('GRID',           (0,0), (-1,-1), 0.5, colors.grey),
                    ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f5f5')]),
                    ('FONTSIZE',       (0,0), (-1,-1), 9),
                ]))
                elements.append(t)
            doc.build(elements)
            QMessageBox.information(self, "Export réussi", f"PDF sauvegardé :\n{path}")
        except ImportError:
            QMessageBox.warning(self, "Dépendance manquante",
                                "reportlab n'est pas installé.\npip install reportlab")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export PDF", str(e))

    def _export_excel(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export Excel", "emploi_du_temps.xlsx", "Excel (*.xlsx)"
        )
        if not path:
            return
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Emploi du temps"
            ws.append(["Date", "Début", "Fin", "Activité", "Type", "Salle", "Enseignant", "α"])
            for col in range(1, 9):
                ws.cell(1, col).font = Font(bold=True, color="FFFFFF")
                ws.cell(1, col).fill = PatternFill(fill_type="solid", fgColor="1a73e8")
            for s in sorted(self._get_current_slots(),
                            key=lambda x: (str(x.get('date', '')), x.get('start', ''))):
                ws.append([str(s.get('date', '')), s.get('start', ''), s.get('end', ''),
                           s.get('activity', ''), s.get('activity_type', ''),
                           s.get('room', ''), s.get('teacher', ''), s.get('alpha', '')])
            wb.save(path)
            QMessageBox.information(self, "Export réussi", f"Excel sauvegardé :\n{path}")
        except ImportError:
            QMessageBox.warning(self, "Dépendance manquante", "pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export Excel", str(e))

    def _export_ical(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export iCal", "emploi_du_temps.ics", "iCal (*.ics)"
        )
        if not path:
            return
        try:
            lines = ["BEGIN:VCALENDAR", "VERSION:2.0",
                     "PRODID:-//Pfair Scheduler//UC9//FR",
                     "CALSCALE:GREGORIAN", "METHOD:PUBLISH"]
            for s in self._get_current_slots():
                try:
                    d = s.get('date')
                    d_str   = d.strftime('%Y%m%d') if hasattr(d, 'strftime') else str(d).replace('-', '')
                    start_t = s.get('start', '08:00').replace(':', '')
                    end_t   = s.get('end',   '10:00').replace(':', '')
                    lines += ["BEGIN:VEVENT",
                              f"DTSTART:{d_str}T{start_t}00",
                              f"DTEND:{d_str}T{end_t}00",
                              f"SUMMARY:{s.get('activity', 'Cours')} [{s.get('activity_type', '')}]",
                              f"LOCATION:{s.get('room', '')}",
                              f"DESCRIPTION:Enseignant: {s.get('teacher', '')}\\n"
                              f"Cohorte: {s.get('cohort', '')}\\nα={s.get('alpha', '')}",
                              "END:VEVENT"]
                except Exception:
                    pass
            lines.append("END:VCALENDAR")
            with open(path, 'w', encoding='utf-8') as f:
                f.write('\r\n'.join(lines))
            QMessageBox.information(self, "Export réussi", f"iCal sauvegardé :\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export iCal", str(e))

    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Export CSV", "emploi_du_temps.csv", "CSV (*.csv)"
        )
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(["Date", "Début", "Fin", "Activité", "Type", "Salle", "Enseignant", "α"])
                for s in sorted(self._get_current_slots(),
                                key=lambda x: (str(x.get('date', '')), x.get('start', ''))):
                    writer.writerow([str(s.get('date', '')), s.get('start', ''), s.get('end', ''),
                                     s.get('activity', ''), s.get('activity_type', ''),
                                     s.get('room', ''), s.get('teacher', ''), s.get('alpha', '')])
            QMessageBox.information(self, "Export réussi", f"CSV sauvegardé :\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Erreur export CSV", str(e))

    # ══════════════════════════════════════════════════════════
    # MODE ÉDITION
    # ══════════════════════════════════════════════════════════

    def _toggle_edit_mode(self, checked):
        self._edit_mode = checked
        if checked:
            self.btn_edit_mode.setText("🔒  Quitter édition")
            QMessageBox.information(
                self, "Mode édition activé",
                "Mode édition activé.\n\n"
                "Faites un clic droit sur un créneau\n"
                "pour le modifier ou le supprimer."
            )
        else:
            self.btn_edit_mode.setText("✏️  Mode édition")
        self._refresh_grid()

    def _on_edit_slot(self, slot_data):
        dialog = SlotEditDialog(slot_data, self._teachers, parent=self)
        if dialog.exec_() != QDialog.Accepted:
            return
        values  = dialog.get_values()
        slot_id = slot_data.get('slot_id')
        if not slot_id:
            QMessageBox.warning(self, "Erreur",
                "Impossible d'identifier ce créneau en base.")
            return
        try:
            from datetime import time as dtime_cls
            from src.database.models import ScheduleSlotModel as SSM
            from sqlalchemy import and_
            session  = db_manager.get_session()
            slot_orm = session.query(SSM).filter_by(id=slot_id).first()
            if not slot_orm:
                QMessageBox.warning(self, "Erreur", "Créneau introuvable en base.")
                return
            sh, sm    = map(int, values['start'].split(':'))
            eh, em    = map(int, values['end'].split(':'))
            new_start = dtime_cls(sh, sm)
            new_end   = dtime_cls(eh, em)
            conflict  = session.query(SSM).filter(
                and_(
                    SSM.id         != slot_id,
                    SSM.date       == values['date'],
                    SSM.teacher_id == values['teacher_id'],
                    SSM.start_time  < new_end,
                    SSM.end_time    > new_start,
                )
            ).first()
            if conflict:
                c_name = conflict.activity.name if conflict.activity else '?'
                QMessageBox.warning(
                    self, "Conflit détecté",
                    f"L'enseignant a déjà un cours à cet horaire :\n"
                    f"{c_name} ({conflict.start_time} - {conflict.end_time})\n\n"
                    "Modification annulée."
                )
                return
            slot_orm.date       = values['date']
            slot_orm.start_time = new_start
            slot_orm.end_time   = new_end
            slot_orm.room       = values['room']
            slot_orm.teacher_id = values['teacher_id']
            session.commit()
            QMessageBox.information(
                self, "Créneau modifié",
                f"Créneau modifié avec succès.\n"
                f"Date : {values['date'].strftime('%d/%m/%Y')}\n"
                f"Horaire : {values['start']} – {values['end']}\n"
                f"Salle : {values['room']}\n"
                f"Enseignant : {values['teacher']}"
            )
            self._reload_all()
        except Exception as exc:
            QMessageBox.critical(self, "Erreur",
                f"Impossible de modifier le créneau :\n{exc}")

    def _on_delete_slot(self, slot_data):
        slot_id  = slot_data.get('slot_id')
        act_name = slot_data.get('activity', '?')
        d        = slot_data.get('date')
        if isinstance(d, str):
            d = datetime.fromisoformat(d).date()
        date_str = d.strftime('%d/%m/%Y') if d else '?'
        reply = QMessageBox.question(
            self, "Confirmer la suppression",
            f"Supprimer ce créneau ?\n\n"
            f"• {act_name}\n"
            f"• {date_str}  {slot_data.get('start','')} – {slot_data.get('end','')}\n"
            f"• Salle : {slot_data.get('room','')}\n\n"
            "Cette action est irréversible.",
            QMessageBox.Yes | QMessageBox.Cancel,
            QMessageBox.Cancel
        )
        if reply != QMessageBox.Yes:
            return
        try:
            session  = db_manager.get_session()
            slot_orm = session.query(ScheduleSlotModel).filter_by(id=slot_id).first()
            if slot_orm:
                session.delete(slot_orm)
                session.commit()
                QMessageBox.information(
                    self, "Créneau supprimé",
                    f"Le créneau '{act_name}' du {date_str} a été supprimé."
                )
                self._reload_all()
            else:
                QMessageBox.warning(self, "Erreur", "Créneau introuvable en base.")
        except Exception as exc:
            QMessageBox.critical(self, "Erreur",
                f"Impossible de supprimer :\n{exc}")

    def refresh_data(self):
        self._reload_all()

    def showEvent(self, event):
        super().showEvent(event)
        self._load_db_data()